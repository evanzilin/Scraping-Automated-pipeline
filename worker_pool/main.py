"""
python main.py setup-db
python main.py append-file --excel-path data.xlsx
python main.py start-workers --worker-count 30 --batch-size 1000
"""

import asyncio
import argparse
import hashlib
import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
from pathlib import Path
import sys
from typing import Any

import httpx
from openpyxl import load_workbook
from psycopg import connect
from psycopg.rows import dict_row

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import check_email_dup as email_pipeline
import main as lead_generator


LOGGER = logging.getLogger("provider")
DEFAULT_DSN = "postgresql://postgres:12345678@localhost:5432/Leads"


SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS provider_files (
    file_name TEXT PRIMARY KEY,
    loaded_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS data (
    id BIGSERIAL PRIMARY KEY,
    source_file TEXT NOT NULL,
    source_row_number BIGINT NOT NULL,
    payload JSONB NOT NULL,
    status TEXT NOT NULL DEFAULT 'pending',
    duplicate_check TEXT,
    email_status TEXT,
    batch_id UUID,
    claimed_by TEXT,
    claimed_at TIMESTAMPTZ,
    acked_at TIMESTAMPTZ,
    created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    CONSTRAINT data_status_check CHECK (status IN ('pending', 'in_progress', 'done')),
    CONSTRAINT data_source_row_unique UNIQUE (source_file, source_row_number)
);

CREATE INDEX IF NOT EXISTS data_pending_idx
    ON data (status, id)
    WHERE status = 'pending';

CREATE INDEX IF NOT EXISTS data_batch_idx
    ON data (batch_id)
    WHERE batch_id IS NOT NULL;

ALTER TABLE data
    ADD COLUMN IF NOT EXISTS duplicate_check TEXT;

ALTER TABLE data
    ADD COLUMN IF NOT EXISTS email_status TEXT;

CREATE INDEX IF NOT EXISTS data_duplicate_check_idx
    ON data (duplicate_check);

CREATE INDEX IF NOT EXISTS data_email_status_idx
    ON data (email_status);

CREATE OR REPLACE FUNCTION claim_batch(p_worker_id TEXT, p_batch_size INTEGER)
RETURNS TABLE (
    batch_id UUID,
    id BIGINT,
    source_file TEXT,
    source_row_number BIGINT,
    payload JSONB
) AS $$
DECLARE
    v_batch_id UUID := gen_random_uuid();
BEGIN
    RETURN QUERY
    WITH picked AS (
        SELECT d.id
        FROM data AS d
        WHERE d.status = 'pending'
        ORDER BY d.id
        FOR UPDATE SKIP LOCKED
        LIMIT p_batch_size
    ),
    updated AS (
        UPDATE data AS d
        SET status = 'in_progress',
            batch_id = v_batch_id,
            claimed_by = p_worker_id,
            claimed_at = NOW()
        FROM picked
        WHERE d.id = picked.id
        RETURNING d.id, d.source_file, d.source_row_number, d.payload
    )
    SELECT v_batch_id, u.id, u.source_file, u.source_row_number, u.payload
    FROM updated AS u
    ORDER BY u.id;
END;
$$ LANGUAGE plpgsql;

CREATE OR REPLACE FUNCTION ack_batch(p_worker_id TEXT, p_batch_id UUID)
RETURNS INTEGER AS $$
DECLARE
    v_count INTEGER;
BEGIN
    UPDATE data
    SET status = 'done',
        acked_at = NOW()
    WHERE batch_id = p_batch_id
      AND claimed_by = p_worker_id
      AND status = 'in_progress';

    GET DIAGNOSTICS v_count = ROW_COUNT;
    RETURN v_count;
END;
$$ LANGUAGE plpgsql;

"""


def get_connection(dsn: str):
    return connect(dsn, autocommit=False, row_factory=dict_row)


def setup_database(dsn: str) -> None:
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute("CREATE EXTENSION IF NOT EXISTS pgcrypto")
            cur.execute(SCHEMA_SQL)
        conn.commit()


def load_excel_headers(excel_path: Path) -> list[str]:
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        workbook.close()
        raise ValueError(f"No active worksheet found in {excel_path}")
    try:
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    finally:
        workbook.close()
    return ["" if value is None else str(value) for value in header_values]


def _initial_duplicate_check(payload: dict[str, Any]) -> str | None:
    value = payload.get("duplicate_check")
    if value is None:
        return None
    text = str(value).strip().lower()
    if text == "dup":
        return "dup"
    if text == "not dup":
        return "not dup"
    return None


def _initial_email_status(payload: dict[str, Any]) -> str | None:
    verification = payload.get(email_pipeline.COL_VERIFICATION)
    if verification is not None:
        verification_text = str(verification).strip().lower()
        if verification_text == email_pipeline.LABEL_VALID_EMAIL.lower():
            return "valid"
        if verification_text == email_pipeline.LABEL_INVALID_EMAIL.lower():
            return "invalid"

    value = payload.get("email_status")
    if value is not None:
        text = str(value).strip().lower()
        if text == "valid":
            return "valid"
        if text == "invalid":
            return "invalid"

    validation_status = payload.get(email_pipeline.COL_STATUS)
    if validation_status is not None:
        status_text = str(validation_status).strip().lower()
        if status_text == email_pipeline.STATUS_PASSED.lower():
            return "valid"
        if status_text == email_pipeline.STATUS_NOT_PASSED.lower():
            return "invalid"

    return None


def append_excel_file(dsn: str, excel_path: Path) -> int:
    excel_path = excel_path.resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    file_name = excel_path.name
    headers = load_excel_headers(excel_path)

    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT 1 FROM provider_files WHERE file_name = %s",
                (file_name,),
            )
            if cur.fetchone():
                conn.rollback()
                LOGGER.info("File %s was already loaded. Skipping.", file_name)
                return 0

        workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
        worksheet = workbook.active
        if worksheet is None:
            workbook.close()
            raise ValueError(f"No active worksheet found in {excel_path}")
        inserted = 0
        try:
            with conn.cursor() as cur:
                for row_number, row_values in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True),
                    start=2,
                ):
                    payload = {
                        headers[index]: row_values[index] if index < len(row_values) else None
                        for index in range(len(headers))
                    }
                    cur.execute(
                        """
                        INSERT INTO data (
                            source_file,
                            source_row_number,
                            payload,
                            duplicate_check,
                            email_status
                        )
                        VALUES (%s, %s, %s::jsonb, %s, %s)
                        ON CONFLICT (source_file, source_row_number) DO NOTHING
                        """,
                        (
                            file_name,
                            row_number,
                            json.dumps(payload, default=str),
                            _initial_duplicate_check(payload),
                            _initial_email_status(payload),
                        ),
                    )
                    inserted += 1

                cur.execute(
                    """
                    INSERT INTO provider_files (file_name)
                    VALUES (%s)
                    ON CONFLICT (file_name) DO NOTHING
                    """,
                    (file_name,),
                )
            conn.commit()
        finally:
            workbook.close()

    LOGGER.info("Loaded %s rows from %s", inserted, file_name)
    return inserted


def claim_batch(dsn: str, worker_id: str, batch_size: int = 1000) -> dict[str, Any]:
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM claim_batch(%s, %s)",
                (worker_id, batch_size),
            )
            rows = cur.fetchall()
        conn.commit()

    if not rows:
        return {"batch_id": None, "rows": []}

    batch_id = str(rows[0]["batch_id"])
    data_rows = [
        {
            "id": row["id"],
            "source_file": row["source_file"],
            "source_row_number": row["source_row_number"],
            "payload": row["payload"],
        }
        for row in rows
    ]
    return {"batch_id": batch_id, "rows": data_rows}


def ack_batch(dsn: str, worker_id: str, batch_id: str) -> int:
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT ack_batch(%s, %s::uuid) AS updated_count",
                (worker_id, batch_id),
            )
            result = cur.fetchone()
        conn.commit()
    return int(result["updated_count"])


def _first_non_blank_string(payload: dict[str, Any], names: tuple[str, ...]) -> str:
    for name in names:
        value = payload.get(name)
        if value is None:
            continue
        text = str(value).strip()
        if text and text.lower() != "nan":
            return text
    return ""


def _lead_payload_from_source_row(source_payload: dict[str, Any]) -> dict[str, Any]:
    lead_payload = dict(source_payload)
    try:
        exported_rows = lead_generator._records_for_not_dup_json_export([source_payload])
        if exported_rows:
            lead_payload.update(exported_rows[0])
    except Exception:
        LOGGER.exception("Failed to normalize source payload into lead format.")

    lead_payload["email"] = str(lead_payload.get("email") or "").strip() or _first_non_blank_string(
        source_payload, email_pipeline._EMAIL_COLUMN_PRIORITY
    )
    lead_payload["first"] = str(lead_payload.get("first") or "").strip() or _first_non_blank_string(
        source_payload, email_pipeline._FIRST_NAME_COLUMN_PRIORITY
    )
    lead_payload["last"] = str(lead_payload.get("last") or "").strip() or _first_non_blank_string(
        source_payload, email_pipeline._LAST_NAME_COLUMN_PRIORITY
    )
    lead_payload["website"] = str(lead_payload.get("website") or "").strip() or _first_non_blank_string(
        source_payload, email_pipeline._WEBSITE_COLUMN_PRIORITY
    )
    if not str(lead_payload.get("business") or "").strip():
        lead_payload["business"] = str(
            source_payload.get("Company Name") or source_payload.get("business") or ""
        ).strip()
    if not str(lead_payload.get("role") or "").strip():
        lead_payload["role"] = str(source_payload.get("Title") or source_payload.get("role") or "").strip()
    if not str(lead_payload.get("company_linkedin") or "").strip():
        lead_payload["company_linkedin"] = str(
            source_payload.get("Company Linkedin Url") or source_payload.get("company_linkedin") or ""
        ).strip()
    return lead_payload


def _validation_reject_reason(result: dict[str, Any]) -> str:
    failed = [
        str(step.get("name") or "").strip()
        for step in result.get("steps", [])
        if isinstance(step, dict) and not bool(step.get("ok"))
    ]
    failed = [name for name in failed if name]
    return ", ".join(failed[:8]) if failed else "invalid_email"


def _set_email_status(payload: dict[str, Any], *, is_valid: bool) -> None:
    payload["email_status"] = "valid" if is_valid else "invalid"


async def _apply_email_validation(source_payload: dict[str, Any]) -> dict[str, Any]:
    lead_payload = _lead_payload_from_source_row(source_payload)
    email = str(lead_payload.get("email") or "").strip()
    first_name = str(lead_payload.get("first") or "").strip()
    last_name = str(lead_payload.get("last") or "").strip()

    if not email:
        lead_payload[email_pipeline.COL_STATUS] = email_pipeline.STATUS_NOT_PASSED
        lead_payload[email_pipeline.COL_VERIFICATION] = email_pipeline.LABEL_INVALID_EMAIL
        lead_payload[email_pipeline.COL_EMAIL_VALIDATION_PASSED] = "false"
        lead_payload[email_pipeline.COL_REASON] = "missing_email"
        lead_payload[email_pipeline.COL_DETAIL] = "No email value found in the row payload."
        lead_payload[email_pipeline.COL_DELIVERABILITY] = "missing_email"
        lead_payload[email_pipeline.COL_MAILBOX_VERIFY] = ""
        _set_email_status(lead_payload, is_valid=False)
        lead_payload["duplicate_check"] = None
        lead_payload["email_hash"] = ""
        return lead_payload

    validation = await email_pipeline.run_single_email_stream(
        email,
        first_name=first_name,
        last_name=last_name,
        website=(str(lead_payload.get("website") or "").strip() or None),
        skip_name_match=not (first_name and last_name),
    )
    is_valid = bool(validation.get("valid"))
    lead_payload["email"] = str(validation.get("email") or email).strip()
    lead_payload[email_pipeline.COL_STATUS] = (
        email_pipeline.STATUS_PASSED if is_valid else email_pipeline.STATUS_NOT_PASSED
    )
    lead_payload[email_pipeline.COL_VERIFICATION] = (
        email_pipeline.LABEL_VALID_EMAIL if is_valid else email_pipeline.LABEL_INVALID_EMAIL
    )
    lead_payload[email_pipeline.COL_EMAIL_VALIDATION_PASSED] = "true" if is_valid else "false"
    lead_payload[email_pipeline.COL_REASON] = "" if is_valid else _validation_reject_reason(validation)
    lead_payload[email_pipeline.COL_DETAIL] = str(validation.get("deliverability_summary") or "").strip()
    lead_payload[email_pipeline.COL_DELIVERABILITY] = str(
        validation.get("deliverability_summary") or ""
    ).strip()
    lead_payload[email_pipeline.COL_MAILBOX_VERIFY] = str(
        validation.get("mailbox_verify_status") or ""
    ).strip()
    _set_email_status(lead_payload, is_valid=is_valid)

    normalized_email = str(validation.get("email") or email).strip().lower()
    lead_payload["email_hash"] = (
        hashlib.sha256(normalized_email.encode("utf-8")).hexdigest() if normalized_email else ""
    )
    lead_payload["duplicate_check"] = None
    lead_payload["duplicate_check_error"] = ""
    return lead_payload


async def _apply_duplicate_check(
    lead_payload: dict[str, Any],
    duplicate_client: httpx.AsyncClient,
) -> dict[str, Any]:
    normalized_email = str(lead_payload.get("email") or "").strip().lower()
    if not normalized_email:
        lead_payload["duplicate_check"] = None
        lead_payload["duplicate_check_error"] = "missing_email_after_validation"
        return lead_payload

    email_hash, duplicate_result = await email_pipeline.is_duplicate_async(
        duplicate_client, normalized_email
    )
    lead_payload["email_hash"] = email_hash
    if duplicate_result == "not dup":
        lead_payload["duplicate_check"] = "not dup"
        lead_payload["duplicate_check_error"] = ""
    elif duplicate_result == "dup":
        lead_payload["duplicate_check"] = "dup"
        lead_payload["duplicate_check_error"] = ""
    else:
        lead_payload["duplicate_check"] = None
        lead_payload["duplicate_check_error"] = str(duplicate_result or "").strip()
    return lead_payload


def _row_is_not_dup(payload: dict[str, Any]) -> bool:
    return str(payload.get("duplicate_check") or "").strip().lower() == "not dup"


def _row_is_terminal(payload: dict[str, Any]) -> bool:
    email_status = str(payload.get("email_status") or "").strip().lower()
    duplicate_check = str(payload.get("duplicate_check") or "").strip().lower()
    return email_status == "invalid" or duplicate_check == "dup"


def _batch_domain_map(
    processed_rows: list[tuple[int, dict[str, Any]]],
) -> dict[str, list[int]]:
    domain_to_row_indices: dict[str, list[int]] = {}
    for idx, (_row_id, payload) in enumerate(processed_rows):
        if not _row_is_not_dup(payload):
            continue
        domain = lead_generator._normalize_domain(str(payload.get("website") or ""))
        if not domain:
            continue
        domain_to_row_indices.setdefault(domain, []).append(idx)
    return domain_to_row_indices


def _cache_not_dup_payloads_for_later_scrape(
    processed_rows: list[tuple[int, dict[str, Any]]],
) -> tuple[list[dict[str, Any]], dict[str, list[int]]]:
    cached_payloads: list[dict[str, Any]] = []
    domain_to_cached_indices: dict[str, list[int]] = {}
    for _row_id, payload in processed_rows:
        if not _row_is_not_dup(payload):
            continue
        domain = lead_generator._normalize_domain(str(payload.get("website") or ""))
        if not domain:
            continue
        cached_payload = dict(payload)
        cached_payloads.append(cached_payload)
        domain_to_cached_indices.setdefault(domain, []).append(len(cached_payloads) - 1)
    return cached_payloads, domain_to_cached_indices


def _merge_batch_enrichment_like_main(
    processed_rows: list[tuple[int, dict[str, Any]]],
    *,
    scrape_api_key: str,
) -> list[tuple[int, dict[str, Any]]]:
    if not scrape_api_key:
        return processed_rows

    cached_payloads, domain_to_cached_indices = _cache_not_dup_payloads_for_later_scrape(
        processed_rows
    )
    if not domain_to_cached_indices:
        return processed_rows

    for domain, cached_indices in domain_to_cached_indices.items():
        try:
            enrichment_result = lead_generator.enrich_company_web_profile(
                domain,
                scrape_api_key,
                is_b2b=True,
                dynamic_first_scrape=False,
                dynamic_about_scrape=False,
                allow_domain_correction=True,
                scrape_linkedin_company_page=True,
                google_search_for_description=True,
                linkedin_scraper_api_key=scrape_api_key,
            )
        except Exception:
            LOGGER.exception("Enrichment failed for %s", domain)
            continue

        scraped_lead = lead_generator.lead_format_json_from_enrichment(enrichment_result)
        for idx in cached_indices:
            cached_payload = cached_payloads[idx]
            record_scraped = lead_generator._supplement_scraped_lead_from_database_profile(
                cached_payload,
                scraped_lead,
                api_key=scrape_api_key,
            )
            cached_payloads[idx] = lead_generator._merge_export_lead_with_scraped(
                cached_payload, record_scraped
            )

    domain_to_row_indices = _batch_domain_map(processed_rows)
    for domain, row_indices in domain_to_row_indices.items():
        cached_indices = domain_to_cached_indices.get(domain, [])
        if not cached_indices:
            continue
        for row_idx, cached_idx in zip(row_indices, cached_indices):
            row_id, payload = processed_rows[row_idx]
            merged_cached_payload = cached_payloads[cached_idx]
            record_scraped = lead_generator._supplement_scraped_lead_from_database_profile(
                payload,
                merged_cached_payload,
                api_key=scrape_api_key,
            )
            processed_rows[row_idx] = (
                row_id,
                lead_generator._merge_export_lead_with_scraped(payload, record_scraped),
            )

    return processed_rows


async def _process_claimed_rows(
    dsn: str,
    worker_id: str,
    batch_id: str,
    rows: list[dict[str, Any]],
) -> tuple[list[tuple[int, dict[str, Any]]], int]:
    scrape_api_key = lead_generator.get_scrape_api_key()
    processed_rows: list[tuple[int, dict[str, Any]]] = []
    completed_count = 0
    async with httpx.AsyncClient() as duplicate_client:
        for row in rows:
            row_id = int(row["id"])
            source_payload = row.get("payload")
            payload_dict = dict(source_payload) if isinstance(source_payload, dict) else {}
            try:
                processed_payload = await _apply_email_validation(payload_dict)
            except Exception as exc:
                LOGGER.exception("Worker row processing failed for row id %s", row_id)
                processed_payload = dict(payload_dict)
                processed_payload["processing_error"] = str(exc)
            processed_row = (row_id, processed_payload)
            processed_rows.append(processed_row)
            _persist_processed_rows(dsn, worker_id, batch_id, [processed_row])
            if str(processed_payload.get("email_status") or "").strip().lower() != "valid":
                _complete_processed_rows(dsn, worker_id, batch_id, [processed_row])
                completed_count += 1
                continue
            try:
                processed_payload = await _apply_duplicate_check(
                    processed_payload, duplicate_client
                )
            except Exception as exc:
                LOGGER.exception("Duplicate check failed for row id %s", row_id)
                processed_payload["duplicate_check"] = None
                processed_payload["duplicate_check_error"] = str(exc)
            processed_row = (row_id, processed_payload)
            processed_rows[-1] = processed_row
            _persist_processed_rows(dsn, worker_id, batch_id, [processed_row])
            if _row_is_terminal(processed_payload):
                _complete_processed_rows(dsn, worker_id, batch_id, [processed_row])
                completed_count += 1
    return (
        _merge_batch_enrichment_like_main(processed_rows, scrape_api_key=scrape_api_key),
        completed_count,
    )


def _persist_processed_rows(
    dsn: str,
    worker_id: str,
    batch_id: str,
    processed_rows: list[tuple[int, dict[str, Any]]],
) -> None:
    if not processed_rows:
        return

    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                UPDATE data
                SET payload = %s::jsonb,
                    duplicate_check = %s,
                    email_status = %s
                WHERE id = %s
                  AND batch_id = %s::uuid
                  AND claimed_by = %s
                  AND status = 'in_progress'
                """,
                [
                    (
                        json.dumps(payload, default=str),
                        _initial_duplicate_check(payload),
                        _initial_email_status(payload),
                        row_id,
                        batch_id,
                        worker_id,
                    )
                    for row_id, payload in processed_rows
                ],
            )
        conn.commit()


def _complete_processed_rows(
    dsn: str,
    worker_id: str,
    batch_id: str,
    processed_rows: list[tuple[int, dict[str, Any]]],
) -> int:
    if not processed_rows:
        return 0

    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                UPDATE data
                SET status = 'done',
                    acked_at = NOW()
                WHERE id = %s
                  AND batch_id = %s::uuid
                  AND claimed_by = %s
                  AND status = 'in_progress'
                """,
                [
                    (row_id, batch_id, worker_id)
                    for row_id, _payload in processed_rows
                ],
            )
            updated_count = cur.rowcount
        conn.commit()
    return int(updated_count or 0)


def process_rows(
    dsn: str,
    worker_id: str,
    batch_id: str,
    rows: list[dict[str, Any]],
) -> int:
    processed_rows, completed_count = asyncio.run(
        _process_claimed_rows(dsn, worker_id, batch_id, rows)
    )
    _persist_processed_rows(dsn, worker_id, batch_id, processed_rows)
    return completed_count


def worker_loop(dsn: str, worker_id: str, batch_size: int = 1000) -> int:
    processed_total = 0

    while True:
        batch = claim_batch(dsn, worker_id, batch_size)
        batch_id = batch["batch_id"]
        rows = batch["rows"]
        if not batch_id:
            LOGGER.info("Worker %s found no more pending rows.", worker_id)
            return processed_total

        completed_early = process_rows(dsn, worker_id, batch_id, rows)
        updated_count = ack_batch(dsn, worker_id, batch_id)
        processed_total += completed_early + updated_count
        LOGGER.info(
            "Worker %s processed batch %s with %s rows.",
            worker_id,
            batch_id,
            completed_early + updated_count,
        )


def start_workers(dsn: str, worker_count: int = 30, batch_size: int = 1000) -> int:
    total_processed = 0
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = [
            executor.submit(worker_loop, dsn, f"worker-{index:02d}", batch_size)
            for index in range(1, worker_count + 1)
        ]
        for future in as_completed(futures):
            total_processed += future.result()

    LOGGER.info(
        "All workers finished. Total acknowledged rows: %s",
        total_processed,
    )
    return total_processed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="PostgreSQL-backed batch provider")
    parser.add_argument(
        "--dsn",
        default=DEFAULT_DSN,
        help="PostgreSQL DSN. Defaults to the built-in local Leads database DSN.",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    subparsers.add_parser("setup-db", help="Create required tables and SQL functions.")

    append_parser = subparsers.add_parser("append-file", help="Append an Excel file into data.")
    append_parser.add_argument("--excel-path", required=True, help="Path to the Excel file.")
    workers_parser = subparsers.add_parser(
        "start-workers",
        help="Start worker threads that claim and acknowledge batches until data is exhausted.",
    )
    workers_parser.add_argument("--worker-count", type=int, default=30)
    workers_parser.add_argument("--batch-size", type=int, default=1000)

    return parser.parse_args()


def main() -> int:
    args = parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    if args.command == "setup-db":
        setup_database(args.dsn)
        return 0

    if args.command == "append-file":
        append_excel_file(args.dsn, Path(args.excel_path))
        return 0

    if args.command == "start-workers":
        start_workers(args.dsn, args.worker_count, args.batch_size)
        return 0

    raise ValueError(f"Unknown command: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())

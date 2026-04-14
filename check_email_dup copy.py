#!/usr/bin/env python3
"""
Single-email validation stream (self-contained).

Workbook CLI (``python check_email_dup.py``): uses ``XLSX_PATH`` / ``SHEET_NAME`` (see constants below).
**Pipeline:** (1) validate each row's email and **write** validation columns; (2) run duplicate check
on passed rows (Subnet API) into ``email_hash`` / ``duplicate_check``; (3) save the workbook in place.
**Resume:** rows with a non-empty ``email_validation_status`` are skipped unless you change the code
to pass ``force=True`` into ``run_workbook_validation``. Each validation run processes up to
``WORKBOOK_MAX_EMAILS`` pending rows (run again for the next batch). Ctrl+C saves during validation.

Dependencies: aiohttp, dnspython, python-whois, disposable-email-domains, openpyxl, httpx, pandas.

SMTP mailbox check: probes ``RCPT TO`` on MX (port 25). ``EMAIL_SMTP_RCPT_VERIFY=0`` disables it.
After a successful RCPT for the real address, a second probe uses a random non-existent local part;
if that is also accepted, the domain is treated as catch-all / accept-all and the email is **invalid**
(``valid`` false, mailbox status ``accept_all``).
``EMAIL_SMTP_CATCHALL_VERIFY=0`` skips the random probe only. ``EMAIL_SMTP_MX_MAX`` (default 8) sets how
many MX hosts to try per probe. ``EMAIL_SMTP_VERIFY_STRICT=1`` is retained for metadata only; inconclusive
SMTP does not mark the row invalid (see ``email_mailbox_verify_status`` = unverifiable).
``EMAIL_SMTP_RCPT_TIMEOUT`` (seconds).

Invalid-row removal: rows with ``email_validation_status`` = False or ``email_verification`` = invalid email
(data rows only; row 1 kept). Use ``--delete-invalid-rows-only`` / ``--prune-invalid-only`` to prune only, or
``--delete-invalid-rows-after`` / ``--prune-invalid`` after a validation run. Or set
``EMAIL_DELETE_INVALID_ROWS_AFTER=1`` to prune after validate without a CLI flag.

**Library use** (install from repo root: ``pip install -e .``)::

    import asyncio
    from pathlib import Path
    import single_email_check as sec

    asyncio.run(
        sec.run_workbook_validation(
            Path("workbook.xlsx"),
            Path("workbook.xlsx"),
            sheet_name="Sheet",
            write_validation_columns=True,
        )
    )

After install, CLI: ``single-email-check --help``. Public names are listed in ``__all__``.
"""

from __future__ import annotations

import asyncio
import hashlib
import os
import uuid
import re
import smtplib
import socket
import ssl
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from zipfile import BadZipFile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast
from urllib.parse import urlparse

import aiohttp
import dns.resolver
import httpx
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import whois
from disposable_email_domains import blocklist as DISPOSABLE_DOMAINS

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_XLSX = SCRIPT_DIR / "database.xlsx"
PREFERRED_SHEET = "apollo-contacts-export (2)"
XLSX_PATH = DEFAULT_INPUT_XLSX
SHEET_NAME = PREFERRED_SHEET

COL_STATUS = "email_validation_status"
# Lowercase "true"/"false" string; aligns with duplicate_check_passed_mask().
COL_EMAIL_VALIDATION_PASSED = "email_validation_passed"
COL_REASON = "email_validation_reject_reason"
COL_VERIFICATION = "email_verification"
COL_DETAIL = "email_validation_detail"
COL_DELIVERABILITY = "email_deliverability_summary"
# Filter in Excel: unverifiable = SMTP inconclusive; accept_all = random RCPT accepted (catch-all).
COL_MAILBOX_VERIFY = "email_mailbox_verify_status"
LABEL_VALID_EMAIL = "valid email"
LABEL_INVALID_EMAIL = "invalid email"
STATUS_PASSED = "True"
STATUS_NOT_PASSED = "False"

# SMTP RCPT probe: set EMAIL_SMTP_RCPT_VERIFY=0 to disable (e.g. outbound port 25 blocked).
# EMAIL_SMTP_VERIFY_STRICT=1: recorded on inconclusive probes (row still valid if heuristics pass).
# EMAIL_SMTP_RCPT_TIMEOUT seconds (default 18).
EMAIL_SMTP_RCPT_VERIFY_ENV = "EMAIL_SMTP_RCPT_VERIFY"
EMAIL_SMTP_VERIFY_STRICT_ENV = "EMAIL_SMTP_VERIFY_STRICT"
EMAIL_SMTP_RCPT_TIMEOUT_ENV = "EMAIL_SMTP_RCPT_TIMEOUT"
# Second SMTP probe (random RCPT) for catch-all hint; independent of the main RCPT toggle.
EMAIL_SMTP_CATCHALL_VERIFY_ENV = "EMAIL_SMTP_CATCHALL_VERIFY"
# How many MX preference-sorted hosts to try per SMTP probe (default 8; max 25).
EMAIL_SMTP_MX_MAX_ENV = "EMAIL_SMTP_MX_MAX"
# Workbook: if set to 1/true/yes/on, same as --delete-invalid-rows-after after each validation run.
EMAIL_DELETE_INVALID_ROWS_AFTER_ENV = "EMAIL_DELETE_INVALID_ROWS_AFTER"

# Subnet lead-search duplicate API (workbook; see --duplicate-check-only / --duplicate-check-after).
DUP_API_LEAD_SEARCH_URL = "https://subnet71.com/api/lead-search"
DUP_CHECK_CONCURRENCY = 20
DUP_CHECK_BATCH_SIZE = 10_000
DUP_CHECK_SAVE_INTERVAL = 10_000

# RCPT outcomes treated as "address accepted at this hop" (RFC 5321; 252 = take/forward attempt).
_SMTP_RCPT_OK_CODES = frozenset({250, 251, 252})
_SMTP_ENV_DISABLED = frozenset({"0", "false", "no", "off"})
_SMTP_STRICT_TRUTHY = frozenset({"1", "true", "yes"})
_WORKBOOK_ENV_TRUTHY = frozenset({"1", "true", "yes", "on"})


def _workbook_env_truthy(name: str) -> bool:
    return os.environ.get(name, "").strip().lower() in _WORKBOOK_ENV_TRUTHY


def _smtp_mx_max_hosts() -> int:
    try:
        n = int(os.environ.get(EMAIL_SMTP_MX_MAX_ENV, "8").strip())
    except ValueError:
        n = 8
    return max(1, min(n, 25))


def _smtp_read_timeout_seconds() -> float:
    try:
        return float(os.environ.get(EMAIL_SMTP_RCPT_TIMEOUT_ENV, "18"))
    except ValueError:
        return 18.0


def _smtp_reply_text(repl: object) -> str:
    if repl is None:
        return ""
    if isinstance(repl, bytes):
        return repl.decode(errors="replace")
    return str(repl)


def _smtp_try_quit(smtp: Optional[smtplib.SMTP]) -> None:
    if smtp is None:
        return
    try:
        smtp.quit()
    except Exception:
        pass

# Email: first non-empty cell among these headers wins (same row).
_EMAIL_COLUMN_PRIORITY: Tuple[str, ...] = (
    "person_email",
    "person_email_analyzed",
    "personal_email",
    "email",
)

# First / last name: first matching header in each list wins (one column per role).
_FIRST_NAME_COLUMN_PRIORITY: Tuple[str, ...] = (
    "person_first_name_unanalyzed",
    "person_first_name",
    "first_name",
    "First_Name",
    "First Name",
    "first",
)
_LAST_NAME_COLUMN_PRIORITY: Tuple[str, ...] = (
    "person_last_name_unanalyzed",
    "person_last_name",
    "last_name",
    "Last_Name",
    "Last Name",
    "last",
)
_WEBSITE_COLUMN_PRIORITY: Tuple[str, ...] = (
    "li_website",
    "sanitized_organization_website",
    "organization_website",
    "website",
    "Website",
)

# ---------------------------------------------------------------------------
# Stage 0–style email string checks (same rules as validator_models/automated_checks)
# ---------------------------------------------------------------------------

_GENERAL_PURPOSE_PREFIXES = (
    "info@",
    "hello@",
    "owner@",
    "ceo@",
    "founder@",
    "contact@",
    "support@",
    "team@",
    "admin@",
    "office@",
    "mail@",
    "connect@",
    "help@",
    "hi@",
    "welcome@",
    "inquiries@",
    "general@",
    "feedback@",
    "ask@",
    "outreach@",
    "communications@",
    "crew@",
    "staff@",
    "community@",
    "reachus@",
    "talk@",
    "service@",
)

_KNOWN_COMPLAINER_LOCAL_PARTS = frozenset(
    {
        "abuse",
        "complaint",
        "complaints",
        "unsubscribe",
        "remove",
        "optout",
        "postmaster",
        "spam",
    }
)

_KNOWN_SPAM_TRAP_LOCAL_PARTS = frozenset(
    {
        "spamtrap",
        "spam.trap",
        "honeypot",
        "trap",
        "noreply",
        "no-reply",
        "do-not-reply",
        "donotreply",
        "test",
        "invalid",
    }
)

_FREE_EMAIL_DOMAINS = frozenset(
    {
        "gmail.com",
        "googlemail.com",
        "yahoo.com",
        "yahoo.co.uk",
        "yahoo.fr",
        "outlook.com",
        "hotmail.com",
        "live.com",
        "msn.com",
        "aol.com",
        "mail.com",
        "protonmail.com",
        "proton.me",
        "icloud.com",
        "me.com",
        "mac.com",
        "zoho.com",
        "yandex.com",
        "gmx.com",
        "mail.ru",
    }
)


def extract_root_domain(website: str) -> str:
    if not website:
        return ""
    w = website.strip()
    if w.startswith(("http://", "https://")):
        domain = urlparse(w).netloc
    else:
        domain = w.strip("/")
    if domain.startswith("www."):
        domain = domain[4:]
    return domain


def _normalize_email_address(raw: Optional[str]) -> str:
    """Lowercase; strip Excel/Unicode junk (ZWSP, BOM, NBSP, fullwidth punctuation) that breaks syntax checks."""
    if raw is None:
        return ""
    s = str(raw)
    s = s.replace("\xa0", " ").replace("\u202f", " ")
    for ch in ("\ufeff", "\u200b", "\u200c", "\u200d", "\u2060"):
        s = s.replace(ch, "")
    s = re.sub(r"[\u1680\u2000-\u200a\u205f\u3000]", " ", s)
    s = unicodedata.normalize("NFKC", s)
    return s.strip().lower()


def check_email_regex(email: str) -> bool:
    email = _normalize_email_address(email)
    if not email:
        return False
    try:
        pattern_ascii = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
        pattern_unicode = r"^[\w._%+-]+@[\w.-]+\.[a-zA-Z]{2,}$"
        ok = bool(re.match(pattern_ascii, email)) or bool(
            re.match(pattern_unicode, email, re.UNICODE)
        )
        if not ok or "+" in email.split("@", 1)[0]:
            return False
        return True
    except Exception:
        return False


def check_name_email_match(email: str, first_name: str, last_name: str) -> bool:
    email = _normalize_email_address(email)
    if not email:
        return False
    first_name = (first_name or "").strip()
    last_name = (last_name or "").strip()
    if not first_name or not last_name:
        return False
    local_part = email.split("@")[0].lower() if "@" in email else email.lower()
    first_normalized = re.sub(r"[^a-z0-9]", "", first_name.lower())
    last_normalized = re.sub(r"[^a-z0-9]", "", last_name.lower())
    local_normalized = re.sub(r"[^a-z0-9]", "", local_part)
    MIN_NAME_MATCH_LENGTH = 3
    patterns = []
    if len(first_normalized) >= MIN_NAME_MATCH_LENGTH:
        patterns.append(first_normalized)
    if len(last_normalized) >= MIN_NAME_MATCH_LENGTH:
        patterns.append(last_normalized)
    patterns.append(f"{first_normalized}{last_normalized}")
    if first_normalized:
        patterns.append(f"{first_normalized[0]}{last_normalized}")
        patterns.append(f"{last_normalized}{first_normalized[0]}")
    patterns = [p for p in patterns if p and len(p) >= MIN_NAME_MATCH_LENGTH]
    name_match = any(pattern in local_normalized for pattern in patterns)
    if not name_match and len(local_normalized) >= MIN_NAME_MATCH_LENGTH:
        if len(first_normalized) >= len(local_normalized) and first_normalized.startswith(
            local_normalized
        ):
            name_match = True
        if (
            not name_match
            and len(last_normalized) >= len(local_normalized)
            and last_normalized.startswith(local_normalized)
        ):
            name_match = True
        if not name_match:
            for length in range(MIN_NAME_MATCH_LENGTH, min(len(first_normalized) + 1, 7)):
                name_prefix = first_normalized[:length]
                if name_prefix == local_normalized or name_prefix in local_normalized:
                    name_match = True
                    break
            if not name_match:
                for length in range(MIN_NAME_MATCH_LENGTH, min(len(last_normalized) + 1, 7)):
                    name_prefix = last_normalized[:length]
                    if name_prefix == local_normalized or name_prefix in local_normalized:
                        name_match = True
                        break
    return bool(name_match)


def check_general_purpose_email(email: str) -> bool:
    el = _normalize_email_address(email)
    if not el:
        return False
    return not any(el.startswith(p) for p in _GENERAL_PURPOSE_PREFIXES)


def check_free_email_domain(email: str) -> bool:
    email = _normalize_email_address(email)
    if not email:
        return False
    if "@" not in email:
        return True
    domain = email.split("@", 1)[1]
    return domain not in _FREE_EMAIL_DOMAINS


def is_disposable_email(email: str) -> bool:
    email = _normalize_email_address(email)
    if not email or "@" not in email:
        return False
    domain = email.split("@", 1)[1]
    return domain in DISPOSABLE_DOMAINS


def _not_disposable_pass(email: str) -> bool:
    """Promise.all helper: pass = not disposable."""
    return not is_disposable_email(email)


def _name_match_for_gather(
    email: str, first_name: str, last_name: str, skip_name_match: bool
) -> bool:
    if skip_name_match:
        return True
    fn, ln = (first_name or "").strip(), (last_name or "").strip()
    if not fn or not ln:
        return False
    return check_name_email_match(email, fn, ln)


def _local_part(email: str) -> str:
    em = _normalize_email_address(email)
    if "@" not in em:
        return ""
    return em.split("@", 1)[0].strip()


def spam_trap_risk_signal(email: str) -> Tuple[bool, str]:
    """Heuristic risk signal for likely trap/non-human mailboxes (non-blocking)."""
    lp = _local_part(email)
    if not lp:
        return False, ""
    lp_norm = re.sub(r"[^a-z0-9._-]", "", lp.lower())
    if lp_norm in _KNOWN_SPAM_TRAP_LOCAL_PARTS:
        return True, f"local_part_exact:{lp_norm}"
    for token in ("spamtrap", "honeypot", "trap", "noreply", "donotreply", "invalid", "test"):
        if token in lp_norm:
            return True, f"local_part_contains:{token}"
    if re.fullmatch(r"test\d{0,4}", lp_norm):
        return True, "local_part_pattern:testN"
    return False, ""


def known_complainer_risk_signal(email: str) -> Tuple[bool, str]:
    """Heuristic risk signal for complaint/abuse role inboxes (non-blocking)."""
    lp = _local_part(email)
    if not lp:
        return False, ""
    lp_norm = re.sub(r"[^a-z0-9._-]", "", lp.lower())
    if lp_norm in _KNOWN_COMPLAINER_LOCAL_PARTS:
        return True, f"local_part_exact:{lp_norm}"
    for token in ("abuse", "complaint", "unsubscribe", "remove", "optout", "postmaster", "spam"):
        if token in lp_norm:
            return True, f"local_part_contains:{token}"
    return False, ""


def email_domain(email: str) -> str:
    email = _normalize_email_address(email)
    if "@" not in email:
        return ""
    return email.split("@", 1)[1]


# ---------------------------------------------------------------------------
# DNS / WHOIS / HTTP (mirrors checks_email.py behaviour)
# ---------------------------------------------------------------------------


async def check_mx_record(domain: str) -> bool:
    """
    True if DNS returns at least one MX record for ``domain`` (no A/AAAA fallback).

    For the main stream, ``domain`` is the **email domain** (part after ``@``).
    """
    d = str(domain).strip()
    if not d:
        return False
    try:
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, lambda: dns.resolver.resolve(d, "MX"))
        return True
    except Exception:
        return False


def _smtp_rcpt_mailbox_check(email: str) -> Tuple[bool, Dict[str, Any]]:
    """Probe recipient via SMTP ``RCPT TO`` against MX hosts.

    Returns ``(False, meta)`` when a server answers with **5xx** for ``RCPT TO`` (mailbox rejected).

    Returns ``(True, meta)`` on **250/251**, when the probe is skipped/disabled, when no MX answers,
    or when the probe is **inconclusive** (timeouts / no definitive RCPT). Returns ``(False, meta)``
    only when **5xx** on ``RCPT TO`` (mailbox rejected). **Catch-all** domains that accept every
    RCPT still look \"valid\" here.

    ``meta["smtp_requirements"]`` records which layers succeeded (MX DNS → TCP :25 → EHLO → MAIL → RCPT).

    Disable entirely with ``EMAIL_SMTP_RCPT_VERIFY=0`` if outbound port 25 is blocked.
    """
    meta: Dict[str, Any] = {}
    req: Dict[str, Any] = {
        "mx_dns_resolved": False,
        "mx_host_count": 0,
        "tcp_port_25_connected": False,
        "smtp_ehlo_completed": False,
        "smtp_mail_from_accepted": False,
        "rcpt_to_definitive": False,
    }

    def _attach_req() -> None:
        meta["smtp_requirements"] = dict(req)

    off = os.environ.get(EMAIL_SMTP_RCPT_VERIFY_ENV, "1").strip().lower()
    if off in _SMTP_ENV_DISABLED:
        meta["skipped"] = True
        meta["detail"] = "smtp_rcpt_disabled_by_env"
        req["stopped_before"] = "env_disabled"
        _attach_req()
        return True, meta

    timeout = _smtp_read_timeout_seconds()

    email = _normalize_email_address(email)
    if not email or "@" not in email:
        meta["skipped"] = True
        meta["detail"] = "no_email"
        req["stopped_before"] = "no_email"
        _attach_req()
        return True, meta

    domain = email.split("@", 1)[1]
    try:
        answers = dns.resolver.resolve(domain, "MX", lifetime=10)
        mx_hosts = sorted(
            (int(r.preference), str(r.exchange).rstrip(".")) for r in answers
        )
    except Exception as ex:
        meta["skipped"] = True
        meta["detail"] = f"mx_lookup:{type(ex).__name__}"
        req["stopped_before"] = "mx_dns_lookup_failed"
        _attach_req()
        return True, meta

    if not mx_hosts:
        meta["skipped"] = True
        meta["detail"] = "no_mx_records"
        req["stopped_before"] = "no_mx_records"
        _attach_req()
        return True, meta

    req["mx_dns_resolved"] = True
    req["mx_host_count"] = len(mx_hosts)

    mail_from = f"postmaster@{domain}"
    last_err: Optional[str] = None

    mx_limit = _smtp_mx_max_hosts()
    with _SMTP_RCPT_CONCURRENCY:
        for _, host in mx_hosts[:mx_limit]:
            if not host or host == ".":
                continue
            smtp: Optional[smtplib.SMTP] = None
            try:
                smtp = smtplib.SMTP(host=host, port=25, timeout=timeout)
                req["tcp_port_25_connected"] = True
                req["last_mx_host"] = host
                smtp.ehlo()
                req["smtp_ehlo_completed"] = True
                if smtp.has_extn("STARTTLS"):
                    try:
                        smtp.starttls(context=ssl.create_default_context())
                        smtp.ehlo()
                    except Exception:
                        pass
                code, repl = smtp.mail(mail_from)
                if code >= 400:
                    last_err = f"mail_from_{code}"
                    _smtp_try_quit(smtp)
                    continue
                req["smtp_mail_from_accepted"] = True
                code, repl = smtp.rcpt(email)
                repl_s = _smtp_reply_text(repl)
                meta["mx_host"] = host
                meta["smtp_rcpt_code"] = code
                meta["smtp_response"] = repl_s[:400]
                if code in _SMTP_RCPT_OK_CODES:
                    meta["rcpt_accepted"] = True
                    req["rcpt_to_definitive"] = True
                    _attach_req()
                    try:
                        smtp.rset()
                    except Exception:
                        pass
                    _smtp_try_quit(smtp)
                    return True, meta
                if 500 <= code < 600:
                    meta["rcpt_rejected"] = True
                    meta["detail"] = f"mailbox_rejected_smtp_{code}"
                    req["rcpt_to_definitive"] = True
                    _attach_req()
                    _smtp_try_quit(smtp)
                    return False, meta
                last_err = f"rcpt_{code}:{repl_s[:100]}"
                _attach_req()
                _smtp_try_quit(smtp)
            except smtplib.SMTPException as ex:
                last_err = f"{host}:SMTPException:{ex!s}"[:200]
                _attach_req()
                _smtp_try_quit(smtp)
            except (socket.timeout, TimeoutError, ConnectionRefusedError, OSError) as ex:
                last_err = f"{host}:{type(ex).__name__}"
                _attach_req()
                _smtp_try_quit(smtp)
            except Exception as ex:
                last_err = f"{host}:{type(ex).__name__}:{ex!s}"[:200]
                _attach_req()
                _smtp_try_quit(smtp)

    meta["inconclusive"] = True
    meta["detail"] = last_err or "smtp_all_mx_failed"
    if not req.get("rcpt_to_definitive"):
        req["stopped_before"] = (
            "rcpt_non_definitive_or_no_tcp"
            if req.get("tcp_port_25_connected")
            else "tcp_or_smtp_session_failed"
        )
    _attach_req()
    if (
        os.environ.get(EMAIL_SMTP_VERIFY_STRICT_ENV, "").strip().lower()
        in _SMTP_STRICT_TRUTHY
    ):
        meta["strict_inconclusive"] = True
    return True, meta


def _smtp_catch_all_probe(domain: str) -> Dict[str, Any]:
    """Send ``RCPT TO`` for a random almost-certainly-nonexistent address at ``domain``.

    If the MX accepts it (250/251), the domain is **likely** catch-all / accept-all. If it returns
    5xx for the probe, the domain is **unlikely** to be full catch-all at this MX. Inconclusive
    cases set ``catch_all_likely`` to ``None``.

    Does not run when ``EMAIL_SMTP_RCPT_VERIFY=0`` or ``EMAIL_SMTP_CATCHALL_VERIFY=0``.
    """
    meta: Dict[str, Any] = {"domain": domain}
    off_rcpt = os.environ.get(EMAIL_SMTP_RCPT_VERIFY_ENV, "1").strip().lower()
    if off_rcpt in _SMTP_ENV_DISABLED:
        meta["catch_all_likely"] = None
        meta["detail"] = "skipped_smtp_rcpt_disabled"
        return meta
    off_ca = os.environ.get(EMAIL_SMTP_CATCHALL_VERIFY_ENV, "1").strip().lower()
    if off_ca in _SMTP_ENV_DISABLED:
        meta["catch_all_likely"] = None
        meta["detail"] = "catch_all_probe_disabled_by_env"
        return meta

    timeout = _smtp_read_timeout_seconds()

    d = (domain or "").strip().lower()
    if not d:
        meta["catch_all_likely"] = None
        meta["detail"] = "no_domain"
        return meta

    probe_local = f"invalid-catchall-probe-{uuid.uuid4().hex[:20]}"
    probe_email = f"{probe_local}@{d}"
    meta["probe_email"] = probe_email

    try:
        answers = dns.resolver.resolve(d, "MX", lifetime=10)
        mx_hosts = sorted(
            (int(r.preference), str(r.exchange).rstrip(".")) for r in answers
        )
    except Exception as ex:
        meta["catch_all_likely"] = None
        meta["detail"] = f"mx_lookup:{type(ex).__name__}"
        return meta

    if not mx_hosts:
        meta["catch_all_likely"] = None
        meta["detail"] = "no_mx_records"
        return meta

    mail_from = f"postmaster@{d}"
    last_err: Optional[str] = None

    mx_limit = _smtp_mx_max_hosts()
    with _SMTP_RCPT_CONCURRENCY:
        for _, host in mx_hosts[:mx_limit]:
            if not host or host == ".":
                continue
            smtp: Optional[smtplib.SMTP] = None
            try:
                smtp = smtplib.SMTP(host=host, port=25, timeout=timeout)
                smtp.ehlo()
                if smtp.has_extn("STARTTLS"):
                    try:
                        smtp.starttls(context=ssl.create_default_context())
                        smtp.ehlo()
                    except Exception:
                        pass
                code, repl = smtp.mail(mail_from)
                if code >= 400:
                    last_err = f"mail_from_{code}"
                    _smtp_try_quit(smtp)
                    continue
                code, repl = smtp.rcpt(probe_email)
                repl_s = _smtp_reply_text(repl)
                meta["mx_host"] = host
                meta["smtp_rcpt_code"] = code
                meta["smtp_response"] = repl_s[:400]
                if code in _SMTP_RCPT_OK_CODES:
                    meta["catch_all_likely"] = True
                    meta["detail"] = "random_rcpt_accepted_likely_catch_all"
                    try:
                        smtp.rset()
                    except Exception:
                        pass
                    _smtp_try_quit(smtp)
                    return meta
                if 500 <= code < 600:
                    meta["catch_all_likely"] = False
                    meta["detail"] = "random_rcpt_rejected_not_catch_all"
                    _smtp_try_quit(smtp)
                    return meta
                last_err = f"rcpt_{code}:{repl_s[:100]}"
                _smtp_try_quit(smtp)
            except smtplib.SMTPException as ex:
                last_err = f"{host}:SMTPException:{ex!s}"[:200]
                _smtp_try_quit(smtp)
            except (socket.timeout, TimeoutError, ConnectionRefusedError, OSError) as ex:
                last_err = f"{host}:{type(ex).__name__}"
                _smtp_try_quit(smtp)
            except Exception as ex:
                last_err = f"{host}:{type(ex).__name__}:{ex!s}"[:200]
                _smtp_try_quit(smtp)

    meta["catch_all_likely"] = None
    meta["inconclusive"] = True
    meta["detail"] = last_err or "catch_all_probe_inconclusive"
    return meta


def whois_domain_age_ok(domain: str) -> tuple[bool, Dict[str, Any]]:
    """Returns (passes_min_age_7d, whois_meta).

    Uses ``quiet=True`` and a global concurrency limit so parallel workbook runs do not
    open dozens of WHOIS sockets at once. Retries with backoff on failure.

    Note: python-whois whois.whois() does not accept timeout or ignore_socket_errors
    (TypeError); the library uses its own sockets.
    """
    meta: Dict[str, Any] = {"domain": domain}
    last_err: Optional[Exception] = None
    for attempt in range(_WHOIS_RETRIES):
        try:
            with _WHOIS_CONCURRENCY:
                w = whois.whois(domain, quiet=True)
            if isinstance(w, dict):
                registrar = w.get("registrar")
                nameservers = w.get("name_servers")
                creation_raw = w.get("creation_date")
            else:
                registrar = getattr(w, "registrar", None)
                nameservers = getattr(w, "name_servers", None)
                creation_raw = getattr(w, "creation_date", None)
            if isinstance(nameservers, list):
                nameservers = nameservers[:3]
            meta["registrar"] = registrar
            meta["nameservers"] = nameservers
            if not creation_raw:
                meta["error"] = "no_creation_date"
                return False, meta
            creation_date = (
                creation_raw[0] if isinstance(creation_raw, list) else creation_raw
            )
            if getattr(creation_date, "tzinfo", None) is not None:
                creation_date = creation_date.replace(tzinfo=None)
            age_days = (datetime.now() - creation_date).days
            meta["age_days"] = age_days
            meta["creation_date"] = creation_date.isoformat()
            return age_days >= 7, meta
        except Exception as e:
            last_err = e
            if attempt + 1 < _WHOIS_RETRIES:
                time.sleep(0.6 * (2**attempt))
    meta["error"] = str(last_err) if last_err else "whois failed"
    return False, meta


def spf_dmarc_for_domain(email: str) -> Dict[str, Any]:
    """
    TXT at the email's domain (SPF) and ``_dmarc.{domain}`` (DMARC).

    Pass a full **email** address; the domain is taken from the part after ``@``.
    """
    out = {"has_spf": False, "has_dmarc": False, "dmarc_policy_strict": False}
    domain = email_domain(email)
    if not domain:
        return out
    try:
        txt_records = dns.resolver.resolve(domain, "TXT")
        for record in txt_records:
            txt_string = "".join(
                s.decode() if isinstance(s, bytes) else s for s in record.strings
            )
            if "v=spf1" in txt_string.lower():
                out["has_spf"] = True
                break
    except Exception:
        pass
    try:
        dmarc_domain = f"_dmarc.{domain}"
        txt_records = dns.resolver.resolve(dmarc_domain, "TXT")
        for record in txt_records:
            txt_string = "".join(
                s.decode() if isinstance(s, bytes) else s for s in record.strings
            )
            tl = txt_string.lower()
            if "v=dmarc1" in tl:
                out["has_dmarc"] = True
                out["dmarc_policy_strict"] = "p=quarantine" in tl or "p=reject" in tl
                break
    except Exception:
        pass
    return out


async def verify_website_reachable(url_or_domain: str) -> bool:
    if not url_or_domain:
        return False
    company_domain = url_or_domain
    if not company_domain.startswith(("http://", "https://")):
        company_domain = f"https://{company_domain}"
    PASS_STATUS_CODES = {200, 202, 301, 302, 307, 308, 401, 403, 405, 429, 500, 502, 503}
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
    }
    ssl_context = ssl.create_default_context()
    ssl_context.set_ciphers("DEFAULT:@SECLEVEL=1")
    ssl_context.check_hostname = True
    ssl_context.verify_mode = ssl.CERT_REQUIRED
    connector = aiohttp.TCPConnector(ssl=ssl_context)
    http_timeout = aiohttp.ClientTimeout(total=10)
    try:
        async with aiohttp.ClientSession(headers=headers, connector=connector) as session:
            head_status = None
            head_error = None
            try:
                async with session.head(
                    company_domain, timeout=http_timeout, allow_redirects=True
                ) as response:
                    head_status = response.status
                    if head_status in PASS_STATUS_CODES:
                        return True
            except aiohttp.ClientError as e:
                head_error = str(e) or "connection_error"
                if "Header value is too long" in head_error or "Got more than" in head_error:
                    return True
            except asyncio.TimeoutError:
                head_error = "timeout"
            except Exception as e:
                head_error = str(e) or type(e).__name__
            try:
                async with session.get(
                    company_domain, timeout=http_timeout, allow_redirects=True
                ) as response:
                    if response.status in PASS_STATUS_CODES:
                        return True
            except aiohttp.ClientError as e:
                get_error = str(e) or "connection_error"
                if "Header value is too long" in get_error or "Got more than" in get_error:
                    return True
            except asyncio.TimeoutError:
                return False
            except Exception:
                return False
    except Exception:
        return False
    return False


def dnsbl_clean(root_domain: str) -> bool:
    """True if not listed (NXDOMAIN / no blacklist code = clean)."""
    query = f"{root_domain}.dbl.cloudflare.com"

    def lookup() -> bool:
        try:
            answers = dns.resolver.resolve(query, "A")
            for rdata in answers:
                rec = str(rdata)
                if rec.startswith("127.0.0."):
                    return False
            return True
        except dns.resolver.NXDOMAIN:
            return True
        except Exception:
            return True

    return lookup()


def _deliverability_report(
    hard_ok: bool,
    soft_failures: List[str],
    smtp_meta: Optional[Dict[str, Any]] = None,
    catch_all_meta: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[str, Any], str]:
    """Structured scope of what we did / did not verify; short line for Excel."""
    smtp_meta = dict(smtp_meta or {})
    ca = dict(catch_all_meta or {})
    cal = ca.get("catch_all_likely")
    checked = [
        "syntax_regex",
        "mx_dns_records",
        "domain_registration_age_min_7d",
        "https_domain_homepage_reachable",
        "dnsbl_cloudflare_dbl",
        "not_disposable_blocklist",
        "not_free_provider_list",
        "not_role_local_prefix",
        "name_vs_local_part_when_names_present",
        "spf_dmarc_txt_lookup_soft",
    ]
    if not smtp_meta.get("skipped"):
        checked.append("smtp_rcpt_to_probe")
    if smtp_meta.get("rcpt_accepted") or smtp_meta.get("rcpt_rejected"):
        checked.append("smtp_definitive_rcpt")

    not_checked = [
        "catch_all_detection",
        "inbox_placement",
    ]
    if cal is True or cal is False:
        not_checked = [x for x in not_checked if x != "catch_all_detection"]
        checked.append("catch_all_detection")
    if not smtp_meta.get("rcpt_accepted") and not smtp_meta.get("rcpt_rejected"):
        not_checked.insert(0, "smtp_definitive_rcpt")

    smtp_probe = {
        k: smtp_meta.get(k)
        for k in (
            "skipped",
            "detail",
            "mx_host",
            "smtp_rcpt_code",
            "rcpt_accepted",
            "rcpt_rejected",
            "inconclusive",
            "smtp_requirements",
        )
        if k in smtp_meta and smtp_meta.get(k) is not None
    }

    if not hard_ok:
        detail = {
            "tier": "failed_hard_gates",
            "mailbox_recipient_verified": False,
            "smtp_probe": smtp_probe,
            "checked_gates": checked,
            "not_checked": not_checked,
            "caveats": [
                "Fix failing hard checks above; SMTP RCPT is only meaningful after core gates pass.",
            ],
        }
        if ca:
            detail["catch_all_probe"] = {
                k: v
                for k, v in ca.items()
                if k
                in (
                    "catch_all_likely",
                    "detail",
                    "probe_email",
                    "mx_host",
                    "smtp_rcpt_code",
                    "inconclusive",
                )
                and v is not None
            }
        return detail, "Failed core validation gates; mailbox verification not applicable."

    mailbox_verified = bool(smtp_meta.get("rcpt_accepted"))
    if cal is True:
        caveats = [
            "Random SMTP RCPT probe was accepted — domain likely catch-all / accept-all "
            "(RCPT does not prove this mailbox exists).",
        ]
    elif cal is False:
        caveats = [
            "Random SMTP RCPT probe was rejected — domain unlikely to be full catch-all at this MX "
            "(still possible: policy bounces later).",
        ]
    else:
        caveats = [
            "Catch-all domains may accept RCPT for non-existent mailboxes; bounces can still occur later.",
        ]
    if smtp_meta.get("inconclusive"):
        caveats.insert(
            0,
            "SMTP RCPT probe was inconclusive (timeouts/firewall). "
            "The address is not marked invalid for that alone; see mailbox_verify_status. "
            "Set EMAIL_SMTP_RCPT_VERIFY=0 to skip SMTP entirely.",
        )

    if smtp_meta.get("rcpt_rejected"):
        tier = "rejected_at_smtp"
        summary = "SMTP RCPT rejected by mail server (mailbox not accepted)."
        caveats.insert(0, "SMTP returned 5xx for RCPT TO — invalid at that MX.")
    elif cal is True and mailbox_verified:
        tier = "rejected_accept_all_domain"
        summary = (
            "Catch-all / accept-all domain (random RCPT accepted); email marked invalid."
        )
        caveats.insert(0, summary)
    elif mailbox_verified and not soft_failures:
        tier = "heuristic_plus_smtp_rcpt_ok"
        summary = (
            "Core validation passed and at least one MX accepted RCPT TO "
            "(still possible: catch-all / later policy bounces)."
        )
    elif mailbox_verified and soft_failures:
        tier = "core_pass_mailbox_verified_with_risk_signals"
        summary = (
            "Core validation passed and mailbox was accepted by SMTP, but "
            "non-blocking risk signals were detected."
        )
    elif soft_failures and smtp_meta.get("skipped"):
        tier = "core_pass_smtp_skipped_with_risk_signals"
        summary = (
            "Core validation passed; SMTP skipped and non-blocking risk signals detected."
        )
    elif soft_failures:
        tier = "core_pass_smtp_inconclusive_with_risk_signals"
        summary = (
            "Core validation passed; SMTP inconclusive and non-blocking risk signals detected."
        )
    elif smtp_meta.get("skipped"):
        tier = "core_pass_smtp_skipped"
        summary = "Core validation passed; SMTP RCPT skipped — mailbox not confirmed."
    else:
        tier = "core_pass_smtp_inconclusive"
        summary = "Core validation passed; SMTP RCPT inconclusive — mailbox not confirmed."

    detail = {
        "tier": tier,
        "mailbox_recipient_verified": mailbox_verified,
        "soft_failures": list(soft_failures),
        "smtp_probe": smtp_probe,
        "checked_gates": checked,
        "not_checked": not_checked,
        "caveats": caveats,
    }
    if ca:
        detail["catch_all_probe"] = {
            k: v
            for k, v in ca.items()
            if k
            in (
                "catch_all_likely",
                "detail",
                "probe_email",
                "mx_host",
                "smtp_rcpt_code",
                "inconclusive",
            )
            and v is not None
        }
    if smtp_meta.get("smtp_response"):
        detail["smtp_response_preview"] = str(smtp_meta["smtp_response"])[:200]
    return detail, summary


# ---------------------------------------------------------------------------
# One stream
# ---------------------------------------------------------------------------

@dataclass
class StepResult:
    name: str
    ok: bool
    detail: Optional[str] = None
    data: Dict[str, Any] = field(default_factory=dict)


@dataclass
class WorkbookRunSummary:
    """Totals after a workbook run (for run_xlsx_validation-style footer)."""

    passed: int
    failed: int
    skipped_no_email: int
    skipped_blank_row: int
    skipped_already_done: int
    interrupted: bool
    deleted_invalid_rows: int = 0

    @property
    def total_validated(self) -> int:
        return self.passed + self.failed


async def run_single_email_stream(
    email: str,
    *,
    first_name: str = "",
    last_name: str = "",
    website: Optional[str] = None,
    skip_name_match: bool = False,
) -> Dict[str, Any]:
    """All checks run concurrently (``asyncio.gather`` / Promise.all); steps are recorded in display order."""
    _ = website
    email = _normalize_email_address(email)
    steps: List[StepResult] = []

    def add(name: str, ok: bool, detail: Optional[str] = None, **data: Any) -> None:
        steps.append(StepResult(name=name, ok=ok, detail=detail, data=dict(data)))

    dom = email_domain(email)
    rd = extract_root_domain(dom) or dom if dom else ""
    site_for_http = f"https://{dom}" if dom else "https://"

    # Concurrent: dnsbl, website, whois, spf, mx, disposable, free, general, name, regex, smtp rcpt, risk signals.
    out = await asyncio.gather(
        asyncio.to_thread(dnsbl_clean, rd),
        verify_website_reachable(site_for_http),
        asyncio.to_thread(whois_domain_age_ok, rd),
        asyncio.to_thread(spf_dmarc_for_domain, email),
        check_mx_record(dom),
        asyncio.to_thread(_not_disposable_pass, email),
        asyncio.to_thread(check_free_email_domain, email),
        asyncio.to_thread(check_general_purpose_email, email),
        asyncio.to_thread(
            _name_match_for_gather, email, first_name, last_name, skip_name_match
        ),
        asyncio.to_thread(check_email_regex, email),
        asyncio.to_thread(_smtp_rcpt_mailbox_check, email),
        asyncio.to_thread(spam_trap_risk_signal, email),
        asyncio.to_thread(known_complainer_risk_signal, email),
    )
    bl_ok = bool(out[0])
    head_ok = bool(out[1])
    age_ok, whois_meta = cast(Tuple[bool, Dict[str, Any]], out[2])
    spf_dm = cast(Dict[str, Any], out[3])
    mx_ok = bool(out[4])
    nd_ok = bool(out[5])
    free_ok = bool(out[6])
    gp_ok = bool(out[7])
    nm_ok = bool(out[8])
    re_ok = bool(out[9])
    _, smtp_meta = cast(Tuple[bool, Dict[str, Any]], out[10])
    spam_trap_hit, spam_trap_detail = cast(Tuple[bool, str], out[11])
    complainer_hit, complainer_detail = cast(Tuple[bool, str], out[12])
    parse_ok = bool(dom and rd)

    add("email_regex", re_ok)
    if skip_name_match:
        add("name_email_match", nm_ok, detail="skipped")
    else:
        fn, ln = first_name.strip(), last_name.strip()
        if not fn or not ln:
            add(
                "name_email_match",
                False,
                detail="provide --first and --last or use --skip-name-match",
            )
        else:
            add("name_email_match", nm_ok)
    add("general_purpose_email", gp_ok)
    add("free_email_domain", free_ok)
    add("not_disposable", nd_ok)
    if not dom:
        add("email_domain_parse", False, detail="no domain")
    elif not rd:
        add("email_domain_parse", False, detail="could not derive domain from email")
    else:
        add("email_domain_parse", True)
    add("domain_age_7d", age_ok, **whois_meta)
    add("mx_record", mx_ok, domain=dom)
    add("spf_dmarc_info", True, detail="soft", domain=dom or "", **spf_dm)
    add("website_reachable", head_ok, url=site_for_http)
    add("dnsbl_clean", bl_ok, root_domain=rd)
    add(
        "spam_trap_risk",
        not spam_trap_hit,
        detail=spam_trap_detail or None,
        risk_flag=spam_trap_hit,
    )
    add(
        "known_complainer_risk",
        not complainer_hit,
        detail=complainer_detail or None,
        risk_flag=complainer_hit,
    )
    smtp_step_detail = smtp_meta.get("detail")
    if smtp_meta.get("smtp_response") and not smtp_step_detail:
        smtp_step_detail = str(smtp_meta.get("smtp_response"))[:160]
    add(
        "smtp_rcpt_mailbox",
        not smtp_meta.get("rcpt_rejected"),
        detail=smtp_step_detail,
        **{
            k: v
            for k, v in smtp_meta.items()
            if k
            in (
                "mx_host",
                "smtp_rcpt_code",
                "rcpt_accepted",
                "rcpt_rejected",
                "inconclusive",
                "skipped",
                "smtp_requirements",
            )
            and v is not None
        },
    )

    catch_all_meta: Dict[str, Any] = {}
    if smtp_meta.get("rcpt_accepted") and dom:
        catch_all_meta = await asyncio.to_thread(_smtp_catch_all_probe, dom)
    cal = catch_all_meta.get("catch_all_likely")
    if smtp_meta.get("rcpt_accepted") and dom:
        add(
            "smtp_catch_all_probe",
            cal is not True,
            detail=str(catch_all_meta.get("detail") or "") or None,
            catch_all_likely=cal,
            **{
                k: v
                for k, v in catch_all_meta.items()
                if k
                in ("probe_email", "mx_host", "smtp_rcpt_code", "inconclusive")
                and v is not None
            },
        )
    else:
        add(
            "smtp_catch_all_probe",
            True,
            detail="skipped_no_prior_rcpt_accept",
            catch_all_likely=None,
        )

    # Hard gates: realistic minimum for "valid email" classification.
    hard_ok = parse_ok and re_ok and nd_ok and mx_ok
    # Soft gates: useful quality/risk signals, but not hard invalidators.
    soft_failures: List[str] = []
    if not nm_ok:
        soft_failures.append("name_email_match")
    if not gp_ok:
        soft_failures.append("general_purpose_email")
    if not free_ok:
        soft_failures.append("free_email_domain")
    if not age_ok:
        soft_failures.append("domain_age_7d")
    if not head_ok:
        soft_failures.append("website_reachable")
    if not bl_ok:
        soft_failures.append("dnsbl_clean")
    if spam_trap_hit:
        soft_failures.append("spam_trap_risk")
    if complainer_hit:
        soft_failures.append("known_complainer_risk")
    if cal is True:
        soft_failures.append("smtp_catch_all")

    # Invalid when core gates fail, SMTP rejects mailbox, or catch-all / accept-all detected.
    ok = (
        hard_ok
        and not smtp_meta.get("rcpt_rejected")
        and cal is not True
    )
    deliv, deliv_summary = _deliverability_report(
        hard_ok, soft_failures, smtp_meta=smtp_meta, catch_all_meta=catch_all_meta
    )
    add(
        "deliverability_scope",
        True,
        detail=deliv_summary,
        mailbox_recipient_verified=bool(smtp_meta.get("rcpt_accepted")),
        tier=deliv["tier"],
    )
    mb_lbl = _mailbox_verify_status_label(
        hard_ok, smtp_meta, has_email=True, catch_all_likely=cal
    )
    return _result(
        email, ok, steps, deliv, deliv_summary, mailbox_verify_status=mb_lbl
    )


def _result(
    email: str,
    ok: bool,
    steps: List[StepResult],
    deliverability: Dict[str, Any],
    deliverability_summary: str,
    *,
    mailbox_verify_status: str = "",
) -> Dict[str, Any]:
    return {
        "email": email,
        "valid": ok,
        "mailbox_verify_status": mailbox_verify_status,
        "deliverability": deliverability,
        "deliverability_summary": deliverability_summary,
        "steps": [
            {
                "name": s.name,
                "ok": s.ok,
                "detail": s.detail,
                **s.data,
            }
            for s in steps
        ],
    }


def resolve_worksheet(
    wb: openpyxl.Workbook,
    *,
    sheet_name: Optional[str] = None,
) -> Tuple[Any, str]:
    """Pick worksheet by name or default. Returns (worksheet, title)."""
    names = list(wb.sheetnames)
    if sheet_name is not None:
        sn = str(sheet_name).strip()
        if sn not in names:
            raise SystemExit(
                f"Unknown sheet {sn!r}. Sheets in this file:\n"
                + "\n".join(f"  {j + 1}: {n}" for j, n in enumerate(names))
            )
        return wb[sn], sn
    if PREFERRED_SHEET in names:
        return wb[PREFERRED_SHEET], PREFERRED_SHEET
    return wb[names[0]], names[0]


def _row_has_status(ws: Any, row_idx: int, s_col: int) -> bool:
    """True if validation status cell is non-empty (row already processed)."""
    v = ws.cell(row=row_idx, column=s_col).value
    if v is None:
        return False
    return len(str(v).strip()) > 0


def _ensure_dup_validation_columns(ws: Any) -> Tuple[int, int, int]:
    """Ensure status, verification, and string passed columns exist; return 1-based indices."""
    s_col = _scan_row1_header_column(ws, COL_STATUS)
    v_col = _scan_row1_header_column(ws, COL_VERIFICATION)
    p_col = _scan_row1_header_column(ws, COL_EMAIL_VALIDATION_PASSED)

    def _append_header(header: str) -> int:
        n = _effective_header_width(ws) + 1
        ws.cell(1, n).value = header
        return n

    if s_col is None:
        s_col = _append_header(COL_STATUS)
    if v_col is None:
        v_col = _append_header(COL_VERIFICATION)
    if p_col is None:
        p_col = _append_header(COL_EMAIL_VALIDATION_PASSED)
    return s_col, v_col, p_col


def _write_row_validation_triplet(
    ws: Any,
    row_idx: int,
    s_col: int,
    v_col: int,
    p_col: int,
    *,
    valid: Optional[bool],
) -> None:
    """Write validation columns; ``valid is None`` means no email on the row."""
    if valid is None:
        ws.cell(row=row_idx, column=s_col).value = STATUS_NOT_PASSED
        ws.cell(row=row_idx, column=v_col).value = LABEL_INVALID_EMAIL
        ws.cell(row=row_idx, column=p_col).value = "false"
    elif valid:
        ws.cell(row=row_idx, column=s_col).value = STATUS_PASSED
        ws.cell(row=row_idx, column=v_col).value = LABEL_VALID_EMAIL
        ws.cell(row=row_idx, column=p_col).value = "true"
    else:
        ws.cell(row=row_idx, column=s_col).value = STATUS_NOT_PASSED
        ws.cell(row=row_idx, column=v_col).value = LABEL_INVALID_EMAIL
        ws.cell(row=row_idx, column=p_col).value = "false"


# Each run: only this many email rows from the sheet, then save and end.
WORKBOOK_MAX_EMAILS = 10000
DEFAULT_WORKBOOK_WORKERS = 10
DEFAULT_WORKBOOK_CHUNK = 10

# WHOIS: cap parallel TCP connections (align with typical --workers); retries in whois_domain_age_ok.
_WHOIS_CONCURRENCY = threading.Semaphore(10)
_WHOIS_RETRIES = 3
_SMTP_RCPT_CONCURRENCY = threading.Semaphore(8)


def _effective_header_width(ws: Any) -> int:
    """Last column on row 1 that has a non-blank header, plus a short forward scan.

    Some workbooks under-report ``max_column``; scanning avoids appending validation
    columns on top of existing data.
    """
    mc = int(ws.max_column or 1)
    last = 1
    for c in range(1, mc + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip():
            last = max(last, c)
    for c in range(mc + 1, mc + 512):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip():
            last = c
    return max(last, mc, 1)


def _mailbox_verify_status_label(
    hard_ok: bool,
    smtp_meta: Dict[str, Any],
    *,
    has_email: bool,
    catch_all_likely: Optional[bool] = None,
) -> str:
    """Excel filter-friendly mailbox proof: ``unverifiable`` = inconclusive SMTP; ``accept_all`` = catch-all."""
    if not has_email:
        return "no_email"
    if not hard_ok:
        return "not_applicable"
    if smtp_meta.get("rcpt_rejected"):
        return "rejected"
    if (
        catch_all_likely is True
        and smtp_meta.get("rcpt_accepted")
    ):
        return "accept_all"
    if smtp_meta.get("rcpt_accepted"):
        return "verified"
    return "unverifiable"


def _save_workbook_safe(wb: Any, output_path: Path) -> None:
    """Save to a temp file, then atomic ``os.replace``; fall back to direct save if replace fails."""
    dest = Path(output_path).resolve()
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_name(f"{dest.stem}.tmp{dest.suffix}")
    try:
        wb.save(tmp)
    except Exception:
        if tmp.is_file():
            try:
                tmp.unlink(missing_ok=True)
            except OSError:
                pass
        raise
    try:
        os.replace(str(tmp), str(dest))
    except OSError:
        try:
            wb.save(dest)
        except Exception as e:
            raise OSError(
                f"Cannot save {dest}: file may be open in Excel or not writable. "
                f"A complete copy is at {tmp}; close the workbook and retry, or rename that file."
            ) from e
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass


def _scan_row1_header_column(ws: Any, name: str) -> Optional[int]:
    """1-based column index where row 1 equals ``name`` (exact strip match), or None."""
    limit = max(_effective_header_width(ws) + 1, int(ws.max_column or 1) + 256)
    for c in range(1, limit + 1):
        if str(ws.cell(1, c).value or "").strip() == name:
            return c
    return None


def _cell_str(ws: Any, row: int, col0: Optional[int]) -> str:
    if col0 is None:
        return ""
    v = ws.cell(row=row + 1, column=col0 + 1).value
    if v is None:
        return ""
    return str(v).strip()


def _resolve_personal_email_columns(headers: List[str]) -> List[int]:
    """0-based column indices for each present header in ``_EMAIL_COLUMN_PRIORITY`` (no duplicates)."""
    found: List[int] = []
    used: set[int] = set()
    for want in _EMAIL_COLUMN_PRIORITY:
        for i, h in enumerate(headers):
            if h == want and i not in used:
                found.append(i)
                used.add(i)
                break
    return found


def _resolve_single_column(
    headers: List[str], priority: Tuple[str, ...]
) -> Tuple[Optional[int], str]:
    """First 0-based column index whose header matches ``priority`` order; else (None, '')."""
    for want in priority:
        for i, h in enumerate(headers):
            if h == want:
                return i, want
    return None, ""


def _row_personal_email(ws: Any, data_row_0based: int, col_indices: List[int]) -> str:
    """First non-empty cell among priority columns; normalized lowercase for validation."""
    for c0 in col_indices:
        s = _normalize_email_address(_cell_str(ws, data_row_0based, c0))
        if s:
            return s
    return ""


def _row_contact_for_validation(
    ws: Any,
    data_row_0based: int,
    *,
    email_col_indices: List[int],
    first_col0: Optional[int],
    last_col0: Optional[int],
    web_col0: Optional[int],
) -> Tuple[str, str, str, str]:
    """Read the database row in order: first name, last name, email (normalized), website."""
    first = _cell_str(ws, data_row_0based, first_col0)
    last = _cell_str(ws, data_row_0based, last_col0)
    email = (
        _row_personal_email(ws, data_row_0based, email_col_indices)
        if email_col_indices
        else ""
    )
    website = _cell_str(ws, data_row_0based, web_col0)
    return first, last, email, website


def _contact_row_is_empty(
    first: str, last: str, email: str, website: str
) -> bool:
    """True when the row has no first, last, email, or website (ignores other columns).

    Used to skip padded / unused rows that Excel still counts in ``max_row``.
    """
    if (email or "").strip():
        return False
    if (first or "").strip():
        return False
    if (last or "").strip():
        return False
    if (website or "").strip():
        return False
    return True


def _workbook_progress_step(total: int) -> int:
    if total <= 0:
        return 1
    if total <= 25:
        return 1
    return max(1, total // 40)


def _excel_cell_norm_lower(v: Any) -> str:
    """Normalize a worksheet cell for string comparison (status / verification columns)."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    return str(v).strip().lower()


def _row_marked_invalid_email(ws: Any, row_idx: int, s_col: int, v_col: int) -> bool:
    """True if this row was validated as invalid (matches how results are written)."""
    st = _excel_cell_norm_lower(ws.cell(row=row_idx, column=s_col).value)
    ver = _excel_cell_norm_lower(ws.cell(row=row_idx, column=v_col).value)
    if ver == LABEL_INVALID_EMAIL.lower():
        return True
    if st == STATUS_NOT_PASSED.lower():
        return True
    return False


def _delete_invalid_email_rows(ws: Any, s_col: int, v_col: int) -> int:
    """Remove data rows (not row 1) where status is False or verification is invalid email.

    Deletes from highest row index downward so indices stay valid.
    """
    max_row = int(ws.max_row or 1)
    to_delete: List[int] = []
    for r in range(2, max_row + 1):
        if _row_marked_invalid_email(ws, r, s_col, v_col):
            to_delete.append(r)
    for r in sorted(to_delete, reverse=True):
        ws.delete_rows(r, 1)
    return len(to_delete)


def remove_invalid_email_rows_from_workbook(
    input_path: Path,
    output_path: Path,
    *,
    sheet_name: Optional[str] = None,
) -> int:
    """Open workbook, delete invalid-email rows on the chosen sheet, save to ``output_path``."""
    if not input_path.is_file():
        raise SystemExit(f"Input file not found: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=False)
    try:
        ws, _title = resolve_worksheet(wb, sheet_name=sheet_name)
        s_col = _scan_row1_header_column(ws, COL_STATUS)
        v_col = _scan_row1_header_column(ws, COL_VERIFICATION)
        if s_col is None or v_col is None:
            raise SystemExit(
                f"Cannot delete invalid rows: row 1 must contain {COL_STATUS!r} and {COL_VERIFICATION!r}."
            )
        n = _delete_invalid_email_rows(ws, s_col, v_col)
        _save_workbook_safe(wb, output_path)
        return n
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Lead duplicate check (subnet API; one sheet via pandas read/write + openpyxl).
# ---------------------------------------------------------------------------


async def is_duplicate_async(client: httpx.AsyncClient, email: str) -> Tuple[str, str]:
    lead_id = hashlib.sha256(email.strip().lower().encode("utf-8")).hexdigest()
    try:
        resp = await client.get(
            DUP_API_LEAD_SEARCH_URL,
            params={"leadId": lead_id, "limit": 1000},
            timeout=20.0,
        )
        resp.raise_for_status()
        data = resp.json()

        found = False
        if isinstance(data, list):
            found = len(data) > 0
        elif isinstance(data, dict):
            for key in ("results", "data", "items"):
                if isinstance(data.get(key), list):
                    found = len(data[key]) > 0
                    break
            else:
                if isinstance(data.get("count"), int):
                    found = data["count"] > 0
                elif isinstance(data.get("total"), int):
                    found = data["total"] > 0

        return lead_id, "dup" if found else "not dup"
    except Exception as e:
        return lead_id, f"error: {e}"


def prepare_dataframe_for_duplicate_check(df: pd.DataFrame) -> pd.DataFrame:
    """Drop cleaner title rows and ensure ``email_hash`` / ``duplicate_check`` columns."""
    print(f"  Shape before clean filter: {df.shape}", flush=True)
    _title_col = None
    if "person_title" in df.columns:
        _title_col = "person_title"
    elif "Title" in df.columns:
        _title_col = "Title"
    if _title_col is not None:
        cleaner_mask = df[_title_col].fillna("").astype(str).str.contains(
            "clean", case=False, na=False
        )
        removed = int(cleaner_mask.sum())
        df = df[~cleaner_mask].reset_index(drop=True)
        print(f"  Removed {removed} cleaner rows ({_title_col}). Shape now: {df.shape}", flush=True)
    else:
        print("  No person_title / Title column — skipping clean-row filter.", flush=True)

    if "email_hash" not in df.columns:
        df["email_hash"] = ""
    if "duplicate_check" not in df.columns:
        df["duplicate_check"] = ""
    return df


def _duplicate_check_email_series(df: pd.DataFrame) -> pd.Series:
    for col in _EMAIL_COLUMN_PRIORITY:
        if col in df.columns:
            return df[col]
    raise SystemExit(
        "No email column found. Add one of: " + ", ".join(_EMAIL_COLUMN_PRIORITY)
    )


def duplicate_check_passed_mask(df: pd.DataFrame) -> pd.Series:
    """Rows considered validation-passed for duplicate checking."""
    if COL_EMAIL_VALIDATION_PASSED in df.columns:
        s = df[COL_EMAIL_VALIDATION_PASSED].fillna("").astype(str)
        return s.str.strip().str.lower() == "true"
    if COL_STATUS in df.columns:
        s = df[COL_STATUS].fillna("").astype(str)
        return s.str.strip().str.lower() == "true"
    if COL_VERIFICATION in df.columns:
        s = df[COL_VERIFICATION].fillna("").astype(str)
        return s.str.strip().str.lower() == LABEL_VALID_EMAIL.lower()
    print(
        "Warning: no email_validation_passed / email_validation_status / "
        "email_verification column — no rows treated as passed. "
        "Run validation with --write-validation-columns (or use --duplicate-check-after alone; "
        "it enables that automatically).",
        flush=True,
    )
    return pd.Series(False, index=df.index)


def _series_all_blank_cells(s: pd.Series) -> bool:
    """True if every cell is missing, '', or the literal 'nan' string (pandas str dtype)."""
    x = s.fillna("").astype(str).str.strip()
    x = x.replace("nan", "")
    return bool(x.eq("").all())


def _strip_excel_padding_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Drop all-empty columns Excel/pandas label as ``Unnamed`` (used-range / blank-header padding).

    Without this, ``read_excel`` → ``to_excel`` round-trips hundreds of blank columns.
    """
    if df.empty or len(df.columns) == 0:
        return df
    drop_cols: List[Any] = []
    for col in df.columns:
        name = str(col).strip()
        if name and not name.lower().startswith("unnamed"):
            continue
        if _series_all_blank_cells(df[col]):
            drop_cols.append(col)
    if not drop_cols:
        return df
    return df.drop(columns=drop_cols)


def _save_duplicate_check_dataframe(path: str, sheet_name: str, df: pd.DataFrame) -> None:
    """Replace one worksheet only; keep other sheets (plain ``df.to_excel(path)`` wipes the file)."""
    out = _strip_excel_padding_columns(df)
    with pd.ExcelWriter(
        path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        out.to_excel(writer, sheet_name=sheet_name, index=False)


async def check_duplicate_batch(
    emails: List[str],
    indices: List[int],
    df: pd.DataFrame,
    *,
    xlsx_path: str,
    sheet_name: str,
    concurrency: int = DUP_CHECK_CONCURRENCY,
) -> Tuple[int, int, int]:
    """Check emails for duplicates, writing into ``df``; checkpoint save every ``DUP_CHECK_SAVE_INTERVAL``."""
    sem = asyncio.Semaphore(concurrency)
    dup_count = not_dup_count = error_count = 0
    checked = 0
    lock = asyncio.Lock()

    async def worker(idx: int, email: str) -> None:
        nonlocal dup_count, not_dup_count, error_count, checked
        async with sem:
            lead_id, result = await is_duplicate_async(client, email)
            async with lock:
                df.at[idx, "email_hash"] = lead_id
                df.at[idx, "duplicate_check"] = result
                if result == "dup":
                    dup_count += 1
                elif result == "not dup":
                    not_dup_count += 1
                else:
                    error_count += 1
                checked += 1
                if checked % DUP_CHECK_SAVE_INTERVAL == 0:
                    print(f"  [{checked}/{len(emails)}] saving checkpoint...", flush=True)
                    _save_duplicate_check_dataframe(xlsx_path, sheet_name, df)

    async with httpx.AsyncClient() as client:
        await asyncio.gather(*[worker(idx, email) for idx, email in zip(indices, emails)])

    return dup_count, not_dup_count, error_count


async def run_duplicate_check_phase(
    xlsx_path: str,
    *,
    sheet_name: Optional[str] = None,
    batch_size: int = DUP_CHECK_BATCH_SIZE,
    concurrency: int = DUP_CHECK_CONCURRENCY,
) -> Dict[str, Any]:
    """Load sheet, filter cleaners, run duplicate API for passed unchecked rows, save."""
    path = str(xlsx_path)
    sheet = sheet_name if sheet_name is not None else SHEET_NAME
    print("Loading xlsx for duplicate check...", flush=True)
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl", dtype=str)
    df = _strip_excel_padding_columns(df)
    df = prepare_dataframe_for_duplicate_check(df)

    passed_mask = duplicate_check_passed_mask(df)
    unchecked_mask = df["duplicate_check"].fillna("").astype(str).str.strip() == ""
    pending_indices = df[passed_mask & unchecked_mask].index[:batch_size].tolist()
    em = _duplicate_check_email_series(df)
    emails = em.loc[pending_indices].fillna("").astype(str).str.strip().str.lower().tolist()

    already_done = int((passed_mask & ~unchecked_mask).sum())
    print(f"  Already checked: {already_done} | To check now: {len(emails)}", flush=True)

    if not emails:
        print(
            "Nothing to check. All passed emails already have duplicate_check values.",
            flush=True,
        )
        _save_duplicate_check_dataframe(path, sheet, df)
        return {
            "dup": 0,
            "not_dup": 0,
            "errors": 0,
            "checked": 0,
            "already_done": already_done,
        }

    print(
        f"\nChecking {len(emails)} emails for duplicates (concurrency={concurrency})...",
        flush=True,
    )
    dup_count, not_dup_count, error_count = await check_duplicate_batch(
        emails,
        pending_indices,
        df,
        xlsx_path=path,
        sheet_name=sheet,
        concurrency=concurrency,
    )

    print(
        f"\nResults: {not_dup_count} not dup | {dup_count} dup | {error_count} errors",
        flush=True,
    )
    print("\nSaving xlsx...", flush=True)
    _save_duplicate_check_dataframe(path, sheet, df)
    print(f"Saved to {path}", flush=True)
    return {
        "dup": dup_count,
        "not_dup": not_dup_count,
        "errors": error_count,
        "checked": len(emails),
        "already_done": already_done,
    }


def _format_validation_progress_line(
    row_idx: int,
    first: str,
    last: str,
    email: str,
    *,
    valid: Optional[bool],
) -> str:
    """One-line console summary: names, email, VALID / INVALID."""
    fn = (first or "").strip() or "—"
    ln = (last or "").strip() or "—"
    em = (email or "").strip() or "—"
    if len(em) > 72:
        em = em[:69] + "..."
    if valid is None:
        verdict = "INVALID (no email)"
    elif valid:
        verdict = "VALID"
    else:
        verdict = "INVALID"
    return f"  row {row_idx}: {fn} | {ln} | {em} -> {verdict}"


async def run_workbook_validation(
    input_path: Path,
    output_path: Path,
    *,
    sheet_name: Optional[str] = None,
    force: bool = False,
    max_emails: int = WORKBOOK_MAX_EMAILS,
    workers: int = DEFAULT_WORKBOOK_WORKERS,
    chunk: int = DEFAULT_WORKBOOK_CHUNK,
    delete_invalid_rows_after: bool = False,
    write_validation_columns: bool = False,
) -> WorkbookRunSummary:
    """Run ``run_single_email_stream`` for up to ``max_emails`` pending rows, then save.

    If ``write_validation_columns`` is True, write ``email_validation_status``,
    ``email_verification``, and ``email_validation_passed`` (``\"true\"``/``\"false\"``) so
    duplicate checking (``duplicate_check_passed_mask`` / subnet API) can treat rows as passed.

    If ``delete_invalid_rows_after`` is True and the run was not interrupted, rows where
    ``email_validation_status`` is False or ``email_verification`` is invalid email are removed.
    """
    if not input_path.is_file():
        raise SystemExit(f"Input file not found: {input_path}")

    wb: Any = None
    interrupted = False
    try:
        # data_only=False: reliable read/write round-trip (data_only=True can break saves for some books).
        wb = openpyxl.load_workbook(input_path, data_only=False)
    except BadZipFile as e:
        raise SystemExit(
            f"Cannot open {input_path} as a valid .xlsx (corrupt or not an Excel zip).\n"
            f"Re-export from Excel or use a real workbook. ({e})"
        ) from e
    summary: Optional[WorkbookRunSummary] = None
    executor: Optional[ThreadPoolExecutor] = None
    try:
        ws, sheet_title = resolve_worksheet(wb, sheet_name=sheet_name)
        print(f"Using sheet: {sheet_title!r}", flush=True)
        if write_validation_columns:
            _ensure_dup_validation_columns(ws)
        cap = max(1, int(max_emails))
        if force:
            print("Force: re-validating all rows.", flush=True)

        wn = max(1, int(workers))
        ck = max(1, int(chunk))
        executor = ThreadPoolExecutor(max_workers=wn)
        asyncio.get_running_loop().set_default_executor(executor)
        print(
            f"Concurrency: up to {wn} rows at once (async + Semaphore); "
            f"thread pool {wn} for blocking I/O inside each row; batch size {ck}. "
            f"Adjust with --workers / --chunk.",
            flush=True,
        )

        s_col = _scan_row1_header_column(ws, COL_STATUS)
        w_v_col = _scan_row1_header_column(ws, COL_VERIFICATION) if write_validation_columns else None
        w_p_col = _scan_row1_header_column(ws, COL_EMAIL_VALIDATION_PASSED) if write_validation_columns else None
        if not force:
            if s_col is not None:
                print(
                    f"Resume: skipping rows with non-empty {COL_STATUS!r} "
                    f"(column {get_column_letter(s_col)}).",
                    flush=True,
                )
            else:
                print(
                    f"No {COL_STATUS!r} column in row 1 — resume skip disabled.",
                    flush=True,
                )
        if write_validation_columns:
            print(
                f"Writing {COL_STATUS!r}, {COL_VERIFICATION!r}, {COL_EMAIL_VALIDATION_PASSED!r} per row.",
                flush=True,
            )
        else:
            print(
                "Results are not written to the workbook; check console for VALID/INVALID per row.",
                flush=True,
            )

        headers = [
            str(ws.cell(1, c).value or "").strip()
            for c in range(1, (ws.max_column or 1) + 1)
        ]

        def _header_col_letter(col0: Optional[int]) -> str:
            if col0 is None:
                return "—"
            return get_column_letter(col0 + 1)

        c_first, first_hdr = _resolve_single_column(headers, _FIRST_NAME_COLUMN_PRIORITY)
        c_last, last_hdr = _resolve_single_column(headers, _LAST_NAME_COLUMN_PRIORITY)
        c_web, web_hdr = _resolve_single_column(headers, _WEBSITE_COLUMN_PRIORITY)

        email_col_indices = _resolve_personal_email_columns(headers)
        if first_hdr:
            print(
                f"First name: {first_hdr!r} (column {_header_col_letter(c_first)})",
                flush=True,
            )
        else:
            print(
                "First name: no column matched "
                f"{_FIRST_NAME_COLUMN_PRIORITY} — name match uses empty first.",
                flush=True,
            )
        if last_hdr:
            print(
                f"Last name: {last_hdr!r} (column {_header_col_letter(c_last)})",
                flush=True,
            )
        else:
            print(
                "Last name: no column matched "
                f"{_LAST_NAME_COLUMN_PRIORITY} — name match uses empty last.",
                flush=True,
            )
        if email_col_indices:
            email_parts = [
                f"{headers[i]!r} ({_header_col_letter(i)})"
                for i in email_col_indices
            ]
            print(
                "Email: first non-empty among (in order): " + ", ".join(email_parts),
                flush=True,
            )
        else:
            print(
                "Warning: no email column found (add one of: "
                + ", ".join(_EMAIL_COLUMN_PRIORITY)
                + ").",
                flush=True,
            )
        if web_hdr:
            print(
                f"Website (optional): {web_hdr!r} (column {_header_col_letter(c_web)})",
                flush=True,
            )

        max_row = ws.max_row or 1
        sem = asyncio.Semaphore(wn)

        async def validate_row(
            row_idx: int, email: str, first: str, last: str, website: str
        ) -> Tuple[int, str, str, Dict[str, Any]]:
            async with sem:
                web = website.strip() or None
                skip_nm = not (first.strip() and last.strip())
                out = await run_single_email_stream(
                    email,
                    first_name=first,
                    last_name=last,
                    website=web,
                    skip_name_match=skip_nm,
                )
                return row_idx, first, last, out

        skipped_done = 0
        skipped_no_email = 0
        skipped_blank_row = 0
        tasks: List[Tuple[int, str, str, str, str]] = []
        for r in range(1, max_row):
            row_idx = r + 1
            if not force and s_col is not None and _row_has_status(ws, row_idx, s_col):
                skipped_done += 1
                continue
            first, last, em, wu = _row_contact_for_validation(
                ws,
                r,
                email_col_indices=email_col_indices,
                first_col0=c_first,
                last_col0=c_last,
                web_col0=c_web,
            )
            if _contact_row_is_empty(first, last, em, wu):
                skipped_blank_row += 1
                continue
            if not em:
                skipped_no_email += 1
                print(
                    _format_validation_progress_line(
                        row_idx, first, last, "", valid=None
                    ),
                    flush=True,
                )
                if (
                    write_validation_columns
                    and s_col is not None
                    and w_v_col is not None
                    and w_p_col is not None
                ):
                    _write_row_validation_triplet(
                        ws,
                        row_idx,
                        s_col,
                        w_v_col,
                        w_p_col,
                        valid=None,
                    )
                continue
            tasks.append((row_idx, em, first, last, wu))

        tasks = tasks[:cap]
        print(
            f"This run: up to {cap:,} email row(s) to validate (not counting skipped-already-done), "
            "then save and end.",
            flush=True,
        )

        if skipped_done:
            print(f"Skipped (already processed): {skipped_done} rows", flush=True)
        if skipped_blank_row:
            print(
                f"Skipped (blank row — no first/last/email/website): {skipped_blank_row} rows",
                flush=True,
            )

        total = len(tasks)
        if total == 0:
            print("No email rows to validate (empty sheet or all rows already have status).", flush=True)
        else:
            print(
                f"Starting validation: {total} email row(s). "
                f"Each row: first | last | email -> VALID/INVALID (Ctrl+C saves).",
                flush=True,
            )
        prog_step = _workbook_progress_step(total)
        done = 0
        passed_n = 0
        failed_n = 0
        try:
            for i in range(0, total, ck):
                batch = tasks[i : i + ck]
                batch_tasks = [asyncio.create_task(validate_row(*t)) for t in batch]
                for fut in asyncio.as_completed(batch_tasks):
                    res = await fut
                    if isinstance(res, tuple):
                        row_idx, first_n, last_n, out = res
                        ok = bool(out.get("valid"))
                        if ok:
                            passed_n += 1
                        else:
                            failed_n += 1
                        if (
                            write_validation_columns
                            and s_col is not None
                            and w_v_col is not None
                            and w_p_col is not None
                        ):
                            _write_row_validation_triplet(
                                ws,
                                row_idx,
                                s_col,
                                w_v_col,
                                w_p_col,
                                valid=ok,
                            )
                        print(
                            _format_validation_progress_line(
                                row_idx,
                                first_n,
                                last_n,
                                str(out.get("email") or ""),
                                valid=ok,
                            ),
                            flush=True,
                        )
                    else:
                        raise res
                    done += 1
                    if total and (
                        done == 1
                        or done == total
                        or done % prog_step == 0
                    ):
                        pct = 100.0 * done / total
                        print(
                            f"  Progress: {done}/{total} ({pct:.1f}%)",
                            flush=True,
                        )
        except KeyboardInterrupt:
            interrupted = True
            print("\nInterrupted (Ctrl+C). Saving progress...", flush=True)
        deleted_invalid = 0
        if delete_invalid_rows_after and not interrupted:
            s_pr = _scan_row1_header_column(ws, COL_STATUS)
            v_pr = _scan_row1_header_column(ws, COL_VERIFICATION)
            if s_pr is None or v_pr is None:
                print(
                    f"Prune skipped: row 1 must include {COL_STATUS!r} and {COL_VERIFICATION!r}.",
                    flush=True,
                )
            else:
                deleted_invalid = _delete_invalid_email_rows(ws, s_pr, v_pr)
                if deleted_invalid:
                    print(
                        f"Deleted {deleted_invalid} row(s) with invalid email "
                        f"({COL_STATUS}={STATUS_NOT_PASSED!r} or {COL_VERIFICATION}={LABEL_INVALID_EMAIL!r}).",
                        flush=True,
                    )
        summary = WorkbookRunSummary(
            passed=passed_n,
            failed=failed_n,
            skipped_no_email=skipped_no_email,
            skipped_blank_row=skipped_blank_row,
            skipped_already_done=skipped_done,
            interrupted=interrupted,
            deleted_invalid_rows=deleted_invalid,
        )
    finally:
        if executor is not None:
            executor.shutdown(wait=True)
        if wb is not None:
            try:
                _save_workbook_safe(wb, output_path)
            except OSError as e:
                raise SystemExit(str(e).strip() or repr(e)) from e
            wb.close()
            if output_path.resolve() == input_path.resolve():
                print(f"Updated (in place): {output_path}", flush=True)
            else:
                print(f"Saved: {output_path}", flush=True)

    if interrupted:
        raise SystemExit(130)
    assert summary is not None
    return summary


def main() -> None:
    """
    Fixed pipeline (no CLI flags): email verification → duplicate check → save ``XLSX_PATH``.

    Tune paths and limits via module constants (``XLSX_PATH``, ``SHEET_NAME``,
    ``WORKBOOK_MAX_EMAILS``, ``DEFAULT_WORKBOOK_*``, ``DUP_CHECK_*``) or call
    ``run_workbook_validation`` / ``run_duplicate_check_phase`` from code.
    """
    in_path = XLSX_PATH.resolve()
    out_path = in_path
    if not in_path.is_file():
        raise SystemExit(f"Input file not found: {in_path}")

    # Same resolution as ``run_workbook_validation`` (``PREFERRED_SHEET`` else first tab).
    # Duplicate-check used to default to ``SHEET_NAME`` only, which breaks when that name
    # is missing from the file (e.g. workbook has ``Sheet1`` only).
    wb_resolve: Any = None
    try:
        wb_resolve = openpyxl.load_workbook(in_path, data_only=False)
        _, resolved_sheet = resolve_worksheet(wb_resolve, sheet_name=None)
    except BadZipFile as e:
        raise SystemExit(
            f"Cannot open {in_path} as a valid .xlsx (corrupt or not an Excel zip).\n"
            f"Re-export from Excel or use a real workbook. ({e})"
        ) from e
    finally:
        if wb_resolve is not None:
            wb_resolve.close()

    print(
        f"Pipeline: verify → duplicate check → save\n"
        f"Loading: {in_path}  |  Sheet: {resolved_sheet!r}",
        flush=True,
    )

    delete_after = _workbook_env_truthy(EMAIL_DELETE_INVALID_ROWS_AFTER_ENV)
    if delete_after:
        print(
            f"Prune after validate: {EMAIL_DELETE_INVALID_ROWS_AFTER_ENV}=1 "
            "(invalid rows removed at end if run completes).",
            flush=True,
        )

    stats = asyncio.run(
        run_workbook_validation(
            in_path,
            out_path,
            sheet_name=resolved_sheet,
            force=False,
            max_emails=WORKBOOK_MAX_EMAILS,
            workers=DEFAULT_WORKBOOK_WORKERS,
            chunk=DEFAULT_WORKBOOK_CHUNK,
            delete_invalid_rows_after=delete_after,
            write_validation_columns=True,
        )
    )

    dup_stats = asyncio.run(
        run_duplicate_check_phase(
            str(out_path),
            sheet_name=resolved_sheet,
            batch_size=DUP_CHECK_BATCH_SIZE,
            concurrency=DUP_CHECK_CONCURRENCY,
        )
    )
    print(
        f"Duplicate check: dup={dup_stats.get('dup')} not_dup={dup_stats.get('not_dup')} "
        f"errors={dup_stats.get('errors')} checked={dup_stats.get('checked')}",
        flush=True,
    )

    print(
        f"Done. (This run capped at {WORKBOOK_MAX_EMAILS:,} email row(s); workbook already saved.)",
        flush=True,
    )
    print(f"\n{'=' * 50}")
    print(f"Total validated : {stats.total_validated:,}")
    print(f"Passed ({STATUS_PASSED})   : {stats.passed:,}")
    print(f"Failed ({STATUS_NOT_PASSED})  : {stats.failed:,}")
    print(f"Skipped (no email): {stats.skipped_no_email:,}")
    print(f"Skipped (blank row): {stats.skipped_blank_row:,}")
    print(f"Skipped (already processed): {stats.skipped_already_done:,}")
    if stats.deleted_invalid_rows:
        print(f"Deleted invalid rows: {stats.deleted_invalid_rows:,}")


# Public API for ``import single_email_check`` / ``pip install -e .`` (see pyproject.toml).
__all__ = (
    # Single address
    "run_single_email_stream",
    # Workbook validation
    "run_workbook_validation",
    "WorkbookRunSummary",
    "remove_invalid_email_rows_from_workbook",
    "resolve_worksheet",
    # Duplicate check (subnet API)
    "run_duplicate_check_phase",
    "check_duplicate_batch",
    "is_duplicate_async",
    "prepare_dataframe_for_duplicate_check",
    "duplicate_check_passed_mask",
    # Paths / defaults
    "SCRIPT_DIR",
    "XLSX_PATH",
    "SHEET_NAME",
    "PREFERRED_SHEET",
    "WORKBOOK_MAX_EMAILS",
    "DEFAULT_WORKBOOK_WORKERS",
    "DEFAULT_WORKBOOK_CHUNK",
    # Duplicate tuning
    "DUP_API_LEAD_SEARCH_URL",
    "DUP_CHECK_CONCURRENCY",
    "DUP_CHECK_BATCH_SIZE",
    "DUP_CHECK_SAVE_INTERVAL",
    # Column / label constants (workbook + dup mask)
    "COL_STATUS",
    "COL_VERIFICATION",
    "COL_EMAIL_VALIDATION_PASSED",
    "LABEL_VALID_EMAIL",
    "LABEL_INVALID_EMAIL",
    "STATUS_PASSED",
    "STATUS_NOT_PASSED",
    # CLI entry (optional)
    "main",
)


if __name__ == "__main__":
    main()

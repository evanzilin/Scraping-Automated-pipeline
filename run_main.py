"""
python run_main.py setup-db
python run_main.py append-file --excel-path data.xlsx
python run_main.py start-workers --worker-count 30 --batch-size 1000
"""

import asyncio
import argparse
import hashlib
import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
from pathlib import Path
import sys
from typing import Any

import httpx
from openpyxl import load_workbook
from psycopg import connect
from psycopg.rows import dict_row

REPO_ROOT = Path(__file__).resolve().parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(REPO_ROOT / ".env")
    load_dotenv()
except ImportError:
    pass

import check_email_dup as email_pipeline
import enrichment as lead_generator


LOGGER = logging.getLogger("provider")
DEFAULT_DSN = "postgresql://postgres:12345678@localhost:5432/Leads"

# Tables and indexes only (functions applied in separate executes in setup_database).
SCHEMA_TABLES_SQL = """
CREATE TABLE IF NOT EXISTS provider_files (
    file_name TEXT PRIMARY KEY,
    loaded_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS data (
    id BIGSERIAL PRIMARY KEY,
    source_file TEXT NOT NULL,
    source_row_number BIGINT NOT NULL,
    payload JSONB NOT NULL,
    status TEXT NOT NULL DEFAULT 'pending',
    duplicate_check TEXT,
    email_status TEXT,
    batch_id UUID,
    claimed_by TEXT,
    claimed_at TIMESTAMPTZ,
    acked_at TIMESTAMPTZ,
    created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    CONSTRAINT data_status_check CHECK (status IN ('pending', 'in_progress', 'done')),
    CONSTRAINT data_source_row_unique UNIQUE (source_file, source_row_number)
);

CREATE INDEX IF NOT EXISTS data_pending_idx
    ON data (status, id)
    WHERE status = 'pending';

CREATE INDEX IF NOT EXISTS data_batch_idx
    ON data (batch_id)
    WHERE batch_id IS NOT NULL;

ALTER TABLE data
    ADD COLUMN IF NOT EXISTS duplicate_check TEXT;

ALTER TABLE data
    ADD COLUMN IF NOT EXISTS email_status TEXT;

CREATE INDEX IF NOT EXISTS data_duplicate_check_idx
    ON data (duplicate_check);

CREATE INDEX IF NOT EXISTS data_email_status_idx
    ON data (email_status);
"""

# PostgreSQL does not allow CREATE OR REPLACE when OUT parameters change; drop every overload.
DROP_CLAIM_BATCH_OVERLOADS_SQL = """
DO $drop_claim$
DECLARE
    fn regprocedure;
BEGIN
    FOR fn IN
        SELECT p.oid::regprocedure
        FROM pg_catalog.pg_proc p
        JOIN pg_catalog.pg_namespace n ON n.oid = p.pronamespace
        WHERE p.proname = 'claim_batch'
          AND n.nspname = 'public'
    LOOP
        EXECUTE 'DROP FUNCTION IF EXISTS ' || fn::text || ' CASCADE';
    END LOOP;
END;
$drop_claim$;
"""

CLAIM_BATCH_FUNCTION_SQL = """
CREATE FUNCTION claim_batch(p_worker_id TEXT, p_batch_size INTEGER)
RETURNS TABLE (
    batch_id UUID,
    id BIGINT,
    source_file TEXT,
    source_row_number BIGINT,
    payload JSONB,
    duplicate_check TEXT,
    email_status TEXT
) AS $$
DECLARE
    v_batch_id UUID := gen_random_uuid();
BEGIN
    RETURN QUERY
    WITH picked AS (
        SELECT d.id
        FROM data AS d
        WHERE d.status = 'pending'
        ORDER BY d.id
        FOR UPDATE SKIP LOCKED
        LIMIT p_batch_size
    ),
    updated AS (
        UPDATE data AS d
        SET status = 'in_progress',
            batch_id = v_batch_id,
            claimed_by = p_worker_id,
            claimed_at = NOW()
        FROM picked
        WHERE d.id = picked.id
        RETURNING d.id, d.source_file, d.source_row_number, d.payload,
                  d.duplicate_check, d.email_status
    )
    SELECT v_batch_id, u.id, u.source_file, u.source_row_number, u.payload,
           u.duplicate_check, u.email_status
    FROM updated AS u
    ORDER BY u.id;
END;
$$ LANGUAGE plpgsql;
"""

ACK_BATCH_FUNCTION_SQL = """
CREATE OR REPLACE FUNCTION ack_batch(p_worker_id TEXT, p_batch_id UUID)
RETURNS INTEGER AS $$
DECLARE
    v_count INTEGER;
BEGIN
    UPDATE data
    SET status = 'done',
        acked_at = NOW()
    WHERE batch_id = p_batch_id
      AND claimed_by = p_worker_id
      AND status = 'in_progress';

    GET DIAGNOSTICS v_count = ROW_COUNT;
    RETURN v_count;
END;
$$ LANGUAGE plpgsql;
"""


def get_connection(dsn: str):
    return connect(dsn, autocommit=False, row_factory=dict_row)


def setup_database(dsn: str) -> None:
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute("CREATE EXTENSION IF NOT EXISTS pgcrypto")
            cur.execute(SCHEMA_TABLES_SQL)
            cur.execute(DROP_CLAIM_BATCH_OVERLOADS_SQL)
            cur.execute(CLAIM_BATCH_FUNCTION_SQL)
            cur.execute(ACK_BATCH_FUNCTION_SQL)
        conn.commit()


def load_excel_headers(excel_path: Path) -> list[str]:
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        workbook.close()
        raise ValueError(f"No active worksheet found in {excel_path}")
    try:
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    finally:
        workbook.close()
    return ["" if value is None else str(value) for value in header_values]


def _initial_duplicate_check(payload: dict[str, Any]) -> str | None:
    value = payload.get("duplicate_check")
    if value is None:
        return None
    text = str(value).strip().lower()
    if text == "dup":
        return "dup"
    if text == "not dup":
        return "not dup"
    return None


def _initial_email_status(payload: dict[str, Any]) -> str | None:
    verification = payload.get(email_pipeline.COL_VERIFICATION)
    if verification is not None:
        verification_text = str(verification).strip().lower()
        if verification_text == email_pipeline.LABEL_VALID_EMAIL.lower():
            return "valid"
        if verification_text == email_pipeline.LABEL_INVALID_EMAIL.lower():
            return "invalid"

    value = payload.get("email_status")
    if value is not None:
        text = str(value).strip().lower()
        if text == "valid":
            return "valid"
        if text == "invalid":
            return "invalid"

    validation_status = payload.get(email_pipeline.COL_STATUS)
    if validation_status is not None:
        status_text = str(validation_status).strip().lower()
        if status_text == email_pipeline.STATUS_PASSED.lower():
            return "valid"
        if status_text == email_pipeline.STATUS_NOT_PASSED.lower():
            return "invalid"

    return None


def append_excel_file(dsn: str, excel_path: Path) -> int:
    excel_path = excel_path.resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    file_name = excel_path.name
    headers = load_excel_headers(excel_path)

    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT 1 FROM provider_files WHERE file_name = %s",
                (file_name,),
            )
            if cur.fetchone():
                conn.rollback()
                LOGGER.info("File %s was already loaded. Skipping.", file_name)
                return 0

        workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
        worksheet = workbook.active
        if worksheet is None:
            workbook.close()
            raise ValueError(f"No active worksheet found in {excel_path}")
        inserted = 0
        try:
            with conn.cursor() as cur:
                for row_number, row_values in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True),
                    start=2,
                ):
                    payload = {
                        headers[index]: row_values[index] if index < len(row_values) else None
                        for index in range(len(headers))
                    }
                    cur.execute(
                        """
                        INSERT INTO data (
                            source_file,
                            source_row_number,
                            payload,
                            duplicate_check,
                            email_status
                        )
                        VALUES (%s, %s, %s::jsonb, %s, %s)
                        ON CONFLICT (source_file, source_row_number) DO NOTHING
                        """,
                        (
                            file_name,
                            row_number,
                            json.dumps(payload, default=str),
                            _initial_duplicate_check(payload),
                            _initial_email_status(payload),
                        ),
                    )
                    inserted += 1

                cur.execute(
                    """
                    INSERT INTO provider_files (file_name)
                    VALUES (%s)
                    ON CONFLICT (file_name) DO NOTHING
                    """,
                    (file_name,),
                )
            conn.commit()
        finally:
            workbook.close()

    LOGGER.info("Loaded %s rows from %s", inserted, file_name)
    return inserted


def _hydrate_payload_from_db_row(
    payload_dict: dict[str, Any],
    row: dict[str, Any],
) -> dict[str, Any]:
    """Copy ``duplicate_check`` / ``email_status`` from table columns into the working payload dict."""
    dup = row.get("duplicate_check")
    if dup is not None and str(dup).strip() != "":
        payload_dict["duplicate_check"] = str(dup).strip().lower()
    em = row.get("email_status")
    if em is not None and str(em).strip() != "":
        payload_dict["email_status"] = str(em).strip().lower()
    return payload_dict


def claim_batch(dsn: str, worker_id: str, batch_size: int = 1000) -> dict[str, Any]:
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM claim_batch(%s, %s)",
                (worker_id, batch_size),
            )
            rows = cur.fetchall()
        conn.commit()

    if not rows:
        return {"batch_id": None, "rows": []}

    batch_id = str(rows[0]["batch_id"])
    data_rows = [
        {
            "id": row["id"],
            "source_file": row["source_file"],
            "source_row_number": row["source_row_number"],
            "payload": row["payload"],
            "duplicate_check": row.get("duplicate_check"),
            "email_status": row.get("email_status"),
        }
        for row in rows
    ]
    return {"batch_id": batch_id, "rows": data_rows}


def ack_batch(dsn: str, worker_id: str, batch_id: str) -> int:
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT ack_batch(%s, %s::uuid) AS updated_count",
                (worker_id, batch_id),
            )
            result = cur.fetchone()
        conn.commit()
    return int(result["updated_count"])


def _first_non_blank_string(payload: dict[str, Any], names: tuple[str, ...]) -> str:
    for name in names:
        value = payload.get(name)
        if value is None:
            continue
        text = str(value).strip()
        if text and text.lower() != "nan":
            return text
    return ""


def _lead_payload_from_source_row(source_payload: dict[str, Any]) -> dict[str, Any]:
    lead_payload = dict(source_payload)
    try:
        exported_rows = lead_generator._records_for_not_dup_json_export([source_payload])
        if exported_rows:
            lead_payload.update(exported_rows[0])
    except Exception:
        LOGGER.exception("Failed to normalize source payload into lead format.")

    lead_payload["email"] = str(lead_payload.get("email") or "").strip() or _first_non_blank_string(
        source_payload, email_pipeline._EMAIL_COLUMN_PRIORITY
    )
    lead_payload["first"] = str(lead_payload.get("first") or "").strip() or _first_non_blank_string(
        source_payload, email_pipeline._FIRST_NAME_COLUMN_PRIORITY
    )
    lead_payload["last"] = str(lead_payload.get("last") or "").strip() or _first_non_blank_string(
        source_payload, email_pipeline._LAST_NAME_COLUMN_PRIORITY
    )
    lead_payload["website"] = str(lead_payload.get("website") or "").strip() or _first_non_blank_string(
        source_payload, email_pipeline._WEBSITE_COLUMN_PRIORITY
    )
    if not str(lead_payload.get("business") or "").strip():
        lead_payload["business"] = str(
            source_payload.get("Company Name") or source_payload.get("business") or ""
        ).strip()
    if not str(lead_payload.get("role") or "").strip():
        lead_payload["role"] = str(source_payload.get("Title") or source_payload.get("role") or "").strip()
    if not str(lead_payload.get("company_linkedin") or "").strip():
        lead_payload["company_linkedin"] = str(
            source_payload.get("Company Linkedin Url") or source_payload.get("company_linkedin") or ""
        ).strip()
    return lead_payload


def _validation_reject_reason(result: dict[str, Any]) -> str:
    failed = [
        str(step.get("name") or "").strip()
        for step in result.get("steps", [])
        if isinstance(step, dict) and not bool(step.get("ok"))
    ]
    failed = [name for name in failed if name]
    return ", ".join(failed[:8]) if failed else "invalid_email"


def _set_email_status(payload: dict[str, Any], *, is_valid: bool) -> None:
    payload["email_status"] = "valid" if is_valid else "invalid"


async def _apply_email_validation(source_payload: dict[str, Any]) -> dict[str, Any]:
    lead_payload = _lead_payload_from_source_row(source_payload)
    email = str(lead_payload.get("email") or "").strip()
    first_name = str(lead_payload.get("first") or "").strip()
    last_name = str(lead_payload.get("last") or "").strip()

    if not email:
        lead_payload[email_pipeline.COL_STATUS] = email_pipeline.STATUS_NOT_PASSED
        lead_payload[email_pipeline.COL_VERIFICATION] = email_pipeline.LABEL_INVALID_EMAIL
        lead_payload[email_pipeline.COL_EMAIL_VALIDATION_PASSED] = "false"
        lead_payload[email_pipeline.COL_REASON] = "missing_email"
        lead_payload[email_pipeline.COL_DETAIL] = "No email value found in the row payload."
        lead_payload[email_pipeline.COL_DELIVERABILITY] = "missing_email"
        lead_payload[email_pipeline.COL_MAILBOX_VERIFY] = ""
        _set_email_status(lead_payload, is_valid=False)
        lead_payload["duplicate_check"] = None
        lead_payload["email_hash"] = ""
        return lead_payload

    validation = await email_pipeline.run_single_email_stream(
        email,
        first_name=first_name,
        last_name=last_name,
        website=(str(lead_payload.get("website") or "").strip() or None),
        skip_name_match=not (first_name and last_name),
    )
    is_valid = bool(validation.get("valid"))
    lead_payload["email"] = str(validation.get("email") or email).strip()
    lead_payload[email_pipeline.COL_STATUS] = (
        email_pipeline.STATUS_PASSED if is_valid else email_pipeline.STATUS_NOT_PASSED
    )
    lead_payload[email_pipeline.COL_VERIFICATION] = (
        email_pipeline.LABEL_VALID_EMAIL if is_valid else email_pipeline.LABEL_INVALID_EMAIL
    )
    lead_payload[email_pipeline.COL_EMAIL_VALIDATION_PASSED] = "true" if is_valid else "false"
    lead_payload[email_pipeline.COL_REASON] = "" if is_valid else _validation_reject_reason(validation)
    lead_payload[email_pipeline.COL_DETAIL] = str(validation.get("deliverability_summary") or "").strip()
    lead_payload[email_pipeline.COL_DELIVERABILITY] = str(
        validation.get("deliverability_summary") or ""
    ).strip()
    lead_payload[email_pipeline.COL_MAILBOX_VERIFY] = str(
        validation.get("mailbox_verify_status") or ""
    ).strip()
    _set_email_status(lead_payload, is_valid=is_valid)

    normalized_email = str(validation.get("email") or email).strip().lower()
    lead_payload["email_hash"] = (
        hashlib.sha256(normalized_email.encode("utf-8")).hexdigest() if normalized_email else ""
    )
    lead_payload["duplicate_check"] = None
    lead_payload["duplicate_check_error"] = ""
    return lead_payload


async def _apply_duplicate_check(
    lead_payload: dict[str, Any],
    duplicate_client: httpx.AsyncClient,
) -> dict[str, Any]:
    normalized_email = str(lead_payload.get("email") or "").strip().lower()
    if not normalized_email:
        lead_payload["duplicate_check"] = None
        lead_payload["duplicate_check_error"] = "missing_email_after_validation"
        return lead_payload

    email_hash, duplicate_result = await email_pipeline.is_duplicate_async(
        duplicate_client, normalized_email
    )
    lead_payload["email_hash"] = email_hash
    if duplicate_result == "not dup":
        lead_payload["duplicate_check"] = "not dup"
        lead_payload["duplicate_check_error"] = ""
    elif duplicate_result == "dup":
        lead_payload["duplicate_check"] = "dup"
        lead_payload["duplicate_check_error"] = ""
    else:
        lead_payload["duplicate_check"] = None
        lead_payload["duplicate_check_error"] = str(duplicate_result or "").strip()
    return lead_payload


def _row_is_not_dup(payload: dict[str, Any]) -> bool:
    return str(payload.get("duplicate_check") or "").strip().lower() == "not dup"


def _row_is_terminal(payload: dict[str, Any]) -> bool:
    email_status = str(payload.get("email_status") or "").strip().lower()
    duplicate_check = str(payload.get("duplicate_check") or "").strip().lower()
    scrape_status = str(payload.get("scrape_status") or "").strip().lower()
    return (
        email_status == "invalid"
        or duplicate_check == "dup"
        or scrape_status == "skipped"
    )


def _linkedin_url_from_payload(payload: dict[str, Any]) -> str:
    return lead_generator._absolute_http_url(str(payload.get("company_linkedin") or ""))


async def _apply_company_linkedin_reachability_check(
    lead_payload: dict[str, Any],
    scrape_api_key: str,
) -> dict[str, Any]:
    website_input = str(lead_payload.get("website") or "").strip()
    cached_company_linkedin = _linkedin_url_from_payload(lead_payload)
    root_input = lead_generator._normalize_domain(website_input)

    socials: dict[str, str] = {}
    linkedin_url = ""
    company_website_url = lead_generator._absolute_http_url(website_input)
    source_url = lead_generator.PROPRIETARY_SOURCE_URL
    source_type = lead_generator.SOURCE_TYPE_PROPRIETARY_DATABASE
    root = root_input

    if root_input and scrape_api_key:
        company_website_url, homepage_html = lead_generator._fetch_homepage_html(
            scrape_api_key,
            root_input,
            timeout=60.0,
            dynamic=True,
            preferred_url=website_input,
        )
        if not homepage_html:
            fixed = lead_generator._pick_corrected_domain_via_google(
                scrape_api_key,
                root_input,
                timeout=60.0,
                country="us",
            )
            if fixed:
                root = fixed
                company_website_url, homepage_html = lead_generator._fetch_homepage_html(
                    scrape_api_key,
                    root,
                    timeout=60.0,
                    dynamic=True,
                )
        if homepage_html and company_website_url:
            exact_host = lead_generator._host_from_http_url(company_website_url)
            if exact_host:
                root = exact_host
            business_name = lead_generator._extract_business_name(homepage_html)
            home_links = lead_generator._absolute_links(company_website_url, homepage_html)
            lead_generator._merge_social(socials, home_links)
            lead_generator._merge_social_from_raw_html(socials, homepage_html)
            linkedin_url = lead_generator._linkedin_company_url_from_google_search(
                scrape_api_key,
                company_name=business_name,
                domain_root=root,
                timeout=60.0,
                country="us",
            ) or lead_generator._merge_linkedin(None, home_links) or lead_generator._linkedin_from_raw_html(
                homepage_html
            ) or ""

            about_url = lead_generator._pick_about_url(root, company_website_url, home_links)
            if about_url and about_url.rstrip("/") != company_website_url.rstrip("/"):
                about_html = lead_generator._scrape_html(
                    scrape_api_key,
                    about_url,
                    timeout=60.0,
                    dynamic=True,
                )
                if about_html:
                    about_links = lead_generator._absolute_links(about_url, about_html)
                    lead_generator._merge_social(socials, about_links)
                    lead_generator._merge_social_from_raw_html(socials, about_html)
                    if not linkedin_url:
                        linkedin_url = lead_generator._merge_linkedin(
                            linkedin_url, about_links
                        ) or lead_generator._linkedin_from_raw_html(about_html) or ""
                source_url = about_url
                source_type = lead_generator.SOURCE_TYPE_COMPANY_WEBSITE
            else:
                source_url = company_website_url or lead_generator.PROPRIETARY_SOURCE_URL
                source_type = (
                    lead_generator.SOURCE_TYPE_COMPANY_WEBSITE
                    if company_website_url
                    else lead_generator.SOURCE_TYPE_PROPRIETARY_DATABASE
                )

    if not linkedin_url:
        linkedin_url = cached_company_linkedin

    lead_payload["company_linkedin"] = linkedin_url
    if company_website_url:
        lead_payload["website"] = company_website_url
    if socials:
        lead_payload["socials"] = socials
    lead_payload["resolved_domain"] = root
    lead_payload["source_url"] = source_url
    lead_payload["source_type"] = (
        lead_generator.SOURCE_TYPE_COMPANY_WEBSITE
        if str(source_url or "").startswith(("http://", "https://"))
        else lead_generator.SOURCE_TYPE_PROPRIETARY_DATABASE
    )
    if not linkedin_url:
        lead_payload["scrape_status"] = "skipped"
        lead_payload["scrape_skip_reason"] = "missing_company_linkedin"
        lead_payload["company_linkedin_reachable"] = False
        return lead_payload

    is_reachable = await email_pipeline.verify_website_reachable(linkedin_url)
    lead_payload["company_linkedin_reachable"] = bool(is_reachable)
    if not is_reachable:
        lead_payload["scrape_status"] = "skipped"
        lead_payload["scrape_skip_reason"] = "unreachable_company_linkedin"
        return lead_payload

    lead_payload["scrape_status"] = "pending"
    lead_payload["scrape_skip_reason"] = ""
    return lead_payload


def _batch_domain_map(
    processed_rows: list[tuple[int, dict[str, Any]]],
) -> dict[str, list[int]]:
    domain_to_row_indices: dict[str, list[int]] = {}
    for idx, (_row_id, payload) in enumerate(processed_rows):
        if not _row_is_not_dup(payload):
            continue
        if str(payload.get("scrape_status") or "").strip().lower() == "skipped":
            continue
        domain = lead_generator._normalize_domain(str(payload.get("website") or ""))
        if not domain:
            continue
        domain_to_row_indices.setdefault(domain, []).append(idx)
    return domain_to_row_indices


def _cache_not_dup_payloads_for_later_scrape(
    processed_rows: list[tuple[int, dict[str, Any]]],
) -> tuple[list[dict[str, Any]], dict[str, list[int]]]:
    cached_payloads: list[dict[str, Any]] = []
    domain_to_cached_indices: dict[str, list[int]] = {}
    for _row_id, payload in processed_rows:
        if not _row_is_not_dup(payload):
            continue
        if str(payload.get("scrape_status") or "").strip().lower() == "skipped":
            continue
        domain = lead_generator._normalize_domain(str(payload.get("website") or ""))
        if not domain:
            continue
        cached_payload = dict(payload)
        cached_payloads.append(cached_payload)
        domain_to_cached_indices.setdefault(domain, []).append(len(cached_payloads) - 1)
    return cached_payloads, domain_to_cached_indices


def _payload_has_value(payload: dict[str, Any], key: str) -> bool:
    value = payload.get(key)
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (list, dict)):
        return bool(value)
    return True


_STRUCTURED_LEAD_KEYS: tuple[str, ...] = (
    "business",
    "full_name",
    "first",
    "last",
    "email",
    "role",
    "website",
    "industry", 
    "sub_industry",
    "country",
    "state",
    "city",
    "linkedin",
    "company_linkedin",
    "source_url",
    "source_type",
    "description",
    "employee_count",
    "hq_country",
    "hq_state",
    "hq_city",
    "phone_numbers",
    "socials",
)

_STRUCTURED_META_KEYS: tuple[str, ...] = (
    "email_status",
    "duplicate_check",
    "duplicate_check_error",
    "email_hash",
    "scrape_status",
    "scrape_skip_reason",
    "company_linkedin_reachable",
    "resolved_domain",
    "domain_corrected_from",
    email_pipeline.COL_STATUS,
    email_pipeline.COL_VERIFICATION,
    email_pipeline.COL_EMAIL_VALIDATION_PASSED,
    email_pipeline.COL_REASON,
    email_pipeline.COL_DETAIL,
    email_pipeline.COL_DELIVERABILITY,
    email_pipeline.COL_MAILBOX_VERIFY,
)

# Only these keys may appear in persisted ``payload`` JSON and temp lead files (strict whitelist).
# Drops email-validation columns, scrape_status, duplicate_check, email_hash, and any other worker keys.
_LEAD_EXPORT_JSON_KEYS: frozenset[str] = frozenset(
    {
        "skipped",
        "error",
        "business",
        "source_type",
        "website",
        "source_url",
        "industry",
        "sub_industry",
        "description",
        "employee_count",
        "hq_country",
        "hq_state",
        "hq_city",
        "company_linkedin",
        "resolved_domain",
        "domain_corrected_from",
        "socials",
        "full_name",
        "first",
        "last",
        "email",
        "role",
        "seniority",
        "linkedin",
        "country",
        "state",
        "city",
        "phone_numbers",
    }
)


def _lead_public_json_for_export(payload: dict[str, Any]) -> dict[str, Any]:
    """Keep only company + contact lead fields; never emit worker / validation / scrape metadata."""
    return {k: payload[k] for k in _LEAD_EXPORT_JSON_KEYS if k in payload}


def _structured_payload_from_lead_payload(payload: dict[str, Any]) -> dict[str, Any]:
    structured: dict[str, Any] = {}
    for key in _STRUCTURED_LEAD_KEYS + _STRUCTURED_META_KEYS:
        if key in payload:
            structured[key] = payload.get(key)
    return structured


_CONTACT_KEYS_FOR_DB_PAYLOAD: tuple[str, ...] = (
    "full_name",
    "first",
    "last",
    "email",
    "role",
    "seniority",
    "linkedin",
    "country",
    "state",
    "city",
    "phone_numbers",
)


def _database_payload_company_enrichment_shape(
    merged: dict[str, Any],
    *,
    enrichment: lead_generator.CompanyWebEnrichmentResult,
) -> dict[str, Any]:
    """
    Persist payload in the same shape as ``CompanyWebEnrichmentResult.to_dict()`` (see ``enrichment.py``),
    using refined ``merged`` values plus enrichment flags (e.g. LinkedIn page scraped).

    Contact fields and pipeline metadata keys are appended for in-worker logic; the final
    DB JSON and temp files strip metadata via ``_lead_public_json_for_export``. Columns
    ``email_status`` / ``duplicate_check`` are still set from the full dict at persist time.
    """
    social_raw = merged.get("socials")
    social: dict[str, str] = {}
    if isinstance(social_raw, dict):
        for k, v in social_raw.items():
            if isinstance(v, str) and v.strip():
                social[str(k)] = v.strip()

    website_v = str(merged.get("website") or "").strip()
    if website_v:
        website_v = lead_generator._absolute_http_url(website_v)

    result_obj = lead_generator.CompanyWebEnrichmentResult(
        skipped=False,
        source_type=str(merged.get("source_type") or "").strip() or None,
        company_website_url=website_v or None,
        company_business_name=str(merged.get("business") or "").strip() or None,
        linkedin_url=str(merged.get("company_linkedin") or "").strip() or None,
        source_url=str(merged.get("source_url") or "").strip() or None,
        social=social,
        industry=str(merged.get("industry") or "").strip() or None,
        sub_industry=str(merged.get("sub_industry") or "").strip() or None,
        resolved_domain=str(merged.get("resolved_domain") or "").strip() or None,
        domain_corrected_from=str(merged.get("domain_corrected_from") or "").strip() or None,
        linkedin_company_page_scraped=enrichment.linkedin_company_page_scraped,
        employee_count=str(merged.get("employee_count") or "").strip() or None,
        hq_country=str(merged.get("hq_country") or "").strip() or None,
        hq_state=str(merged.get("hq_state") or "").strip() or None,
        hq_city=str(merged.get("hq_city") or "").strip() or None,
        description=str(merged.get("description") or "").strip() or None,
        error=None,
    )
    out: dict[str, Any] = dict(result_obj.to_dict())
    for key in _CONTACT_KEYS_FOR_DB_PAYLOAD + _STRUCTURED_META_KEYS:
        if key in merged:
            out[key] = merged[key]
    return out


def _normalize_source_fields(payload: dict[str, Any]) -> dict[str, Any]:
    source_url = str(payload.get("source_url") or "").strip()
    if source_url.startswith(("http://", "https://")):
        payload["source_url"] = source_url
        payload["source_type"] = lead_generator.SOURCE_TYPE_COMPANY_WEBSITE
    else:
        payload["source_url"] = lead_generator.PROPRIETARY_SOURCE_URL
        payload["source_type"] = lead_generator.SOURCE_TYPE_PROPRIETARY_DATABASE
    return payload


def _normalize_industry_pair_in_scope(
    industry: str,
    sub_industry: str,
) -> tuple[str, str]:
    ind_text = (industry or "").strip()
    sub_text = (sub_industry or "").strip()
    if not ind_text or not sub_text:
        return ind_text, sub_text
    try:
        from lead_cache_json.industry_normalizer import normalize_industry_pair

        ind_out, sub_out, _reason = normalize_industry_pair(ind_text, sub_text)
        if ind_out and sub_out:
            return ind_out, sub_out
    except ImportError:
        pass
    return ind_text, sub_text


def _apply_cached_keyword_fallback(
    base_payload: dict[str, Any],
    merged_payload: dict[str, Any],
) -> dict[str, Any]:
    industry_seed = str(
        merged_payload.get("industry") or base_payload.get("industry") or ""
    ).strip()
    if industry_seed and not _payload_has_value(merged_payload, "industry"):
        merged_payload["industry"] = industry_seed

    description_v = str(merged_payload.get("description") or "").strip()
    sub_out = str(merged_payload.get("sub_industry") or "").strip()
    if not sub_out and industry_seed:
        guessed_sub = lead_generator._pick_sub_industry_from_database_hints(
            industry_seed,
            str(base_payload.get("_db_sub_industry_hint") or ""),
            str(base_payload.get("_db_keywords") or ""),
            description_v,
        )
        if guessed_sub:
            _, normalized_sub = _normalize_industry_pair_in_scope(
                str(merged_payload.get("industry") or industry_seed),
                guessed_sub,
            )
            if normalized_sub:
                merged_payload["sub_industry"] = normalized_sub
    return merged_payload


def _refine_payload_industry_scope(
    base_payload: dict[str, Any],
    merged_payload: dict[str, Any],
) -> dict[str, Any]:
    industry_v = str(
        merged_payload.get("industry") or base_payload.get("industry") or ""
    ).strip()
    sub_industry_v = str(merged_payload.get("sub_industry") or "").strip()
    description_v = str(merged_payload.get("description") or "").strip()

    refine_candidates: list[str] = []
    for candidate in (
        industry_v,
        sub_industry_v,
        str(base_payload.get("_db_sub_industry_hint") or "").strip(),
        str(base_payload.get("_db_keywords") or "").strip(),
    ):
        if candidate:
            refine_candidates.append(candidate)

    refined_industry, refined_sub_industry = lead_generator._refine_industry_sub_with_taxonomy(
        industry_v or None,
        sub_industry_v or None,
        refine_candidates,
    )
    if refined_industry:
        merged_payload["industry"] = refined_industry
    if refined_sub_industry:
        merged_payload["sub_industry"] = refined_sub_industry

    if not str(merged_payload.get("sub_industry") or "").strip() and description_v:
        guessed_sub = lead_generator._pick_sub_industry_from_database_hints(
            str(merged_payload.get("industry") or industry_v or "").strip(),
            str(base_payload.get("_db_sub_industry_hint") or ""),
            str(base_payload.get("_db_keywords") or ""),
            description_v,
        )
        if guessed_sub:
            normalized_industry, normalized_sub_industry = _normalize_industry_pair_in_scope(
                str(merged_payload.get("industry") or industry_v or "").strip(),
                guessed_sub,
            )
            if normalized_industry:
                merged_payload["industry"] = normalized_industry
            if normalized_sub_industry:
                merged_payload["sub_industry"] = normalized_sub_industry

    try:
        from lead_cache_json.industry_normalizer import coerce_industry_sub_to_taxonomy_scope

        coerced_industry, coerced_sub_industry = coerce_industry_sub_to_taxonomy_scope(
            str(merged_payload.get("industry") or "").strip() or None,
            str(merged_payload.get("sub_industry") or "").strip() or None,
        )
        if coerced_industry:
            merged_payload["industry"] = coerced_industry
        if coerced_sub_industry:
            merged_payload["sub_industry"] = coerced_sub_industry
    except ImportError:
        pass

    return merged_payload


def _normalize_employee_count_scope(merged_payload: dict[str, Any]) -> dict[str, Any]:
    raw_employee_count = str(merged_payload.get("employee_count") or "").strip()
    if not raw_employee_count:
        return merged_payload
    canonical = lead_generator._canonical_employee_count_label(raw_employee_count)
    if canonical:
        merged_payload["employee_count"] = canonical
    return merged_payload


def _apply_linkedin_profile_fallback(
    base_payload: dict[str, Any],
    merged_payload: dict[str, Any],
    *,
    api_key: str,
    timeout: float = 60.0,
) -> dict[str, Any]:
    linkedin_url = lead_generator._absolute_http_url(
        str(
            merged_payload.get("company_linkedin")
            or base_payload.get("company_linkedin")
            or ""
        )
    )
    if not linkedin_url or not (api_key or "").strip():
        return merged_payload

    li_slug = lead_generator._linkedin_company_slug_from_url(linkedin_url)
    if not li_slug:
        return merged_payload

    li_data, li_alt = lead_generator._fetch_scrapingdog_linkedin_company_profile(
        api_key,
        li_slug,
        timeout=timeout,
    )
    if not li_data and not li_alt:
        return merged_payload

    industry_seed = str(
        merged_payload.get("industry") or base_payload.get("industry") or ""
    ).strip()
    li_candidates = lead_generator._linkedin_profile_api_industry_candidates(
        li_data, li_alt
    )
    raw_ind, _ = lead_generator._merge_raw_industry_from_candidates(
        li_candidates,
        industry_seed,
        None,
        assign_sub_industry_from_candidates=False,
    )
    ind_out, _ = lead_generator._refine_industry_sub_with_taxonomy(
        raw_ind,
        None,
        li_candidates,
    )
    if not ind_out and industry_seed:
        ind_out = industry_seed

    description_v = str(merged_payload.get("description") or "").strip()
    if not description_v:
        description_v = str(
            lead_generator._linkedin_profile_api_description(li_data, li_alt) or ""
        ).strip()
        if description_v:
            merged_payload["description"] = description_v

    if ind_out and not _payload_has_value(merged_payload, "industry"):
        merged_payload["industry"] = ind_out

    if not _payload_has_value(merged_payload, "sub_industry") and ind_out:
        sub_out = lead_generator._pick_sub_industry_from_database_hints(
            ind_out,
            str(base_payload.get("_db_sub_industry_hint") or ""),
            str(base_payload.get("_db_keywords") or ""),
            description_v,
        )
        if sub_out:
            normalized_ind, normalized_sub = _normalize_industry_pair_in_scope(
                ind_out, sub_out
            )
            if normalized_ind and not _payload_has_value(merged_payload, "industry"):
                merged_payload["industry"] = normalized_ind
            if normalized_sub:
                merged_payload["sub_industry"] = normalized_sub

    li_api_ec: str | None = None
    li_api_hq: tuple[str | None, str | None, str | None] = (None, None, None)
    for blob in (li_data, li_alt):
        if not blob:
            continue
        ec_p, hq_p = lead_generator._linkedin_profile_api_employee_and_hq(blob)
        if ec_p and not li_api_ec:
            li_api_ec = ec_p
        li_api_hq = lead_generator._merge_hq_triple(li_api_hq, hq_p)

    hq_city_v, hq_state_v, hq_country_v = li_api_hq
    inferred_country, inferred_state = lead_generator._infer_united_states_hq_if_us_state(
        hq_country_v, hq_state_v
    )
    hq_country_v = inferred_country or None
    hq_state_v = inferred_state or None
    if hq_country_v and lead_generator._is_united_states_country(hq_country_v):
        hq_country_v = lead_generator._display_hq_country(hq_country_v)
    if hq_state_v and lead_generator._is_united_states_country(hq_country_v or ""):
        hq_state_v = (
            lead_generator._expand_us_state_abbreviation(hq_state_v.strip())
            or hq_state_v
        )

    merged_payload["company_linkedin"] = linkedin_url
    if li_api_ec and not _payload_has_value(merged_payload, "employee_count"):
        merged_payload["employee_count"] = lead_generator._pick_employee_count(li_api_ec)
    if hq_country_v and not _payload_has_value(merged_payload, "hq_country"):
        merged_payload["hq_country"] = hq_country_v
    if hq_state_v and not _payload_has_value(merged_payload, "hq_state"):
        merged_payload["hq_state"] = hq_state_v
    if hq_city_v and not _payload_has_value(merged_payload, "hq_city"):
        merged_payload["hq_city"] = hq_city_v
    return merged_payload


def _merge_batch_enrichment_like_main(
    processed_rows: list[tuple[int, dict[str, Any]]],
    *,
    scrape_api_key: str,
) -> list[tuple[int, dict[str, Any]]]:
    if not (scrape_api_key or "").strip():
        LOGGER.warning(
            "Scrapingdog API key is missing (set SCRAPE_API_KEY or SCRAPINGDOG_API_KEY in %s or the "
            "environment). Skipping enrich_company_web_profile and domain batch scrape.",
            REPO_ROOT / ".env",
        )
        return processed_rows

    cached_payloads, domain_to_cached_indices = _cache_not_dup_payloads_for_later_scrape(
        processed_rows
    )
    if not domain_to_cached_indices:
        not_dup = sum(1 for _rid, p in processed_rows if _row_is_not_dup(p))
        skipped = sum(
            1
            for _rid, p in processed_rows
            if _row_is_not_dup(p)
            and str(p.get("scrape_status") or "").strip().lower() == "skipped"
        )
        blank_domain = sum(
            1
            for _rid, p in processed_rows
            if _row_is_not_dup(p)
            and str(p.get("scrape_status") or "").strip().lower() != "skipped"
            and not lead_generator._normalize_domain(str(p.get("website") or ""))
        )
        LOGGER.warning(
            "No domains selected for Scrapingdog enrichment (not_dup_rows=%s, "
            "linkedin_skipped=%s, not_dup_missing_website_domain=%s).",
            not_dup,
            skipped,
            blank_domain,
        )
        return processed_rows

    for domain, cached_indices in domain_to_cached_indices.items():
        try:
            enrichment_result = lead_generator.enrich_company_web_profile(
                domain,
                scrape_api_key,
                is_b2b=True,
                dynamic_first_scrape=True,
                dynamic_about_scrape=True,
                allow_domain_correction=True,
                scrape_linkedin_company_page=True,
                google_search_for_description=True,
                linkedin_scraper_api_key=scrape_api_key,
            )
        except Exception:
            LOGGER.exception("Enrichment failed for %s", domain)
            continue

        scraped_lead = lead_generator.lead_format_json_from_enrichment(enrichment_result)
        first_cached = cached_payloads[cached_indices[0]]
        display_supplemented = lead_generator._supplement_scraped_lead_from_database_profile(
            first_cached,
            scraped_lead,
            api_key=scrape_api_key,
        )
        display_enrichment = lead_generator._overlay_enrichment_display_with_lead(
            enrichment_result.to_dict(),
            display_supplemented,
        )
        print(f"\n{'=' * 60}\n[Scrapingdog] domain={domain}\n{'=' * 60}", flush=True)
        if enrichment_result.error:
            print(
                f"Partial result note: {enrichment_result.error}",
                flush=True,
            )
        print(
            "Company profile (enrichment dict, same shape as enrichment.py without --lead-format):",
            flush=True,
        )
        print(json.dumps(display_enrichment, indent=2, ensure_ascii=False), flush=True)
        print("\nLead-format JSON (from enrichment):", flush=True)
        print(json.dumps(scraped_lead, indent=2, ensure_ascii=False), flush=True)

        for idx in cached_indices:
            cached_payload = cached_payloads[idx]
            structured_cached_payload = _structured_payload_from_lead_payload(cached_payload)
            record_scraped = lead_generator._supplement_scraped_lead_from_database_profile(
                cached_payload,
                scraped_lead,
                api_key=scrape_api_key,
            )
            merged_payload = lead_generator._merge_export_lead_with_scraped(
                structured_cached_payload, record_scraped
            )
            merged_payload = _apply_linkedin_profile_fallback(
                cached_payload,
                merged_payload,
                api_key=scrape_api_key,
            )
            merged_payload = _apply_cached_keyword_fallback(
                cached_payload,
                merged_payload,
            )
            merged_payload = _refine_payload_industry_scope(
                cached_payload,
                merged_payload,
            )
            merged_payload = _normalize_employee_count_scope(merged_payload)
            merged_payload = _normalize_source_fields(merged_payload)
            merged_payload["scrape_status"] = "done"
            merged_payload["scrape_skip_reason"] = ""
            cached_payloads[idx] = _database_payload_company_enrichment_shape(
                merged_payload,
                enrichment=enrichment_result,
            )

    domain_to_row_indices = _batch_domain_map(processed_rows)
    for domain, row_indices in domain_to_row_indices.items():
        cached_indices = domain_to_cached_indices.get(domain, [])
        if not cached_indices:
            continue
        for row_idx, cached_idx in zip(row_indices, cached_indices):
            row_id, _payload = processed_rows[row_idx]
            merged_cached_payload = cached_payloads[cached_idx]
            processed_rows[row_idx] = (row_id, merged_cached_payload)

    return processed_rows


async def _process_claimed_rows(
    dsn: str,
    worker_id: str,
    batch_id: str,
    rows: list[dict[str, Any]],
) -> tuple[list[tuple[int, dict[str, Any]]], int]:
    scrape_api_key = lead_generator.get_scrape_api_key()
    processed_rows: list[tuple[int, dict[str, Any]]] = []
    completed_count = 0
    async with httpx.AsyncClient() as duplicate_client:
        for row in rows:
            row_id = int(row["id"])
            source_payload = row.get("payload")
            payload_dict = dict(source_payload) if isinstance(source_payload, dict) else {}
            _hydrate_payload_from_db_row(payload_dict, row)
            try:
                processed_payload = await _apply_email_validation(payload_dict)
            except Exception as exc:
                LOGGER.exception("Worker row processing failed for row id %s", row_id)
                processed_payload = dict(payload_dict)
                processed_payload["processing_error"] = str(exc)
            processed_row = (row_id, processed_payload)
            processed_rows.append(processed_row)
            _persist_email_status_column_only(
                dsn,
                worker_id,
                batch_id,
                row_id,
                email_status_value=(
                    "invalid"
                    if str(processed_payload.get("email_status") or "").strip().lower() != "valid"
                    else _email_status_column_value(processed_payload)
                ),
            )
            if str(processed_payload.get("email_status") or "").strip().lower() != "valid":
                _complete_processed_rows(dsn, worker_id, batch_id, [processed_row])
                completed_count += 1
                continue
            try:
                processed_payload = await _apply_duplicate_check(
                    processed_payload, duplicate_client
                )
            except Exception as exc:
                LOGGER.exception("Duplicate check failed for row id %s", row_id)
                processed_payload["duplicate_check"] = None
                processed_payload["duplicate_check_error"] = str(exc)
            processed_row = (row_id, processed_payload)
            processed_rows[-1] = processed_row
            dup_raw = str(processed_payload.get("duplicate_check") or "").strip().lower()
            if dup_raw == "dup":
                _persist_duplicate_columns_only(
                    dsn,
                    worker_id,
                    batch_id,
                    row_id,
                    email_status_value=_email_status_column_value(processed_payload),
                )
            elif dup_raw == "not dup":
                _persist_not_dup_columns_only(
                    dsn,
                    worker_id,
                    batch_id,
                    row_id,
                    email_status_value=_email_status_column_value(processed_payload),
                )
            else:
                _persist_duplicate_check_unset_columns_only(
                    dsn,
                    worker_id,
                    batch_id,
                    row_id,
                    email_status_value=_email_status_column_value(processed_payload),
                )
            if _row_is_terminal(processed_payload):
                _complete_processed_rows(dsn, worker_id, batch_id, [processed_row])
                completed_count += 1
                continue
            try:
                processed_payload = await _apply_company_linkedin_reachability_check(
                    processed_payload,
                    scrape_api_key,
                )
            except Exception as exc:
                LOGGER.exception(
                    "Company LinkedIn reachability check failed for row id %s", row_id
                )
                processed_payload["scrape_status"] = "skipped"
                processed_payload["scrape_skip_reason"] = str(exc)
                processed_payload["company_linkedin_reachable"] = False
            processed_row = (row_id, processed_payload)
            processed_rows[-1] = processed_row
            if _row_is_terminal(processed_payload):
                _complete_processed_rows(dsn, worker_id, batch_id, [processed_row])
                completed_count += 1
                continue

    merged_rows = _merge_batch_enrichment_like_main(
        processed_rows, scrape_api_key=scrape_api_key
    )
    # Only write ``payload`` after Scrapingdog enrichment produced the final row (see filter).
    _persist_final_scrape_payloads_only(dsn, worker_id, batch_id, merged_rows)
    return (merged_rows, completed_count)


def _email_status_column_value(payload: dict[str, Any]) -> str:
    """Value for ``data.email_status``; must be non-empty for duplicate-only updates."""
    col = _initial_email_status(payload)
    if col in ("valid", "invalid"):
        return col
    raw = str(payload.get("email_status") or "").strip().lower()
    if raw in ("valid", "invalid"):
        return raw
    return "valid"


def _persist_duplicate_columns_only(
    dsn: str,
    worker_id: str,
    batch_id: str,
    row_id: int,
    *,
    email_status_value: str,
) -> None:
    """Mark duplicate in DB without rewriting ``payload`` JSONB (avoids clobbering source data)."""
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE data
                SET duplicate_check = 'dup',
                    email_status = %s
                WHERE id = %s
                  AND batch_id = %s::uuid
                  AND claimed_by = %s
                  AND status = 'in_progress'
                """,
                (email_status_value, row_id, batch_id, worker_id),
            )
        conn.commit()


def _persist_email_status_column_only(
    dsn: str,
    worker_id: str,
    batch_id: str,
    row_id: int,
    *,
    email_status_value: str,
) -> None:
    """Update ``email_status`` only; never touches ``payload`` JSONB."""
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE data
                SET email_status = %s
                WHERE id = %s
                  AND batch_id = %s::uuid
                  AND claimed_by = %s
                  AND status = 'in_progress'
                """,
                (email_status_value, row_id, batch_id, worker_id),
            )
        conn.commit()


def _persist_not_dup_columns_only(
    dsn: str,
    worker_id: str,
    batch_id: str,
    row_id: int,
    *,
    email_status_value: str,
) -> None:
    """Set ``duplicate_check`` to not-dup and refresh ``email_status`` without changing ``payload``."""
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE data
                SET duplicate_check = 'not dup',
                    email_status = %s
                WHERE id = %s
                  AND batch_id = %s::uuid
                  AND claimed_by = %s
                  AND status = 'in_progress'
                """,
                (email_status_value, row_id, batch_id, worker_id),
            )
        conn.commit()


def _persist_duplicate_check_unset_columns_only(
    dsn: str,
    worker_id: str,
    batch_id: str,
    row_id: int,
    *,
    email_status_value: str,
) -> None:
    """Clear ``duplicate_check`` after API error; does not touch ``payload``."""
    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE data
                SET duplicate_check = NULL,
                    email_status = %s
                WHERE id = %s
                  AND batch_id = %s::uuid
                  AND claimed_by = %s
                  AND status = 'in_progress'
                """,
                (email_status_value, row_id, batch_id, worker_id),
            )
        conn.commit()


def _row_should_write_payload_after_scrape(payload: dict[str, Any]) -> bool:
    """True when Scrapingdog enrichment finished and this row is ready for the one payload write."""
    if str(payload.get("email_status") or "").strip().lower() != "valid":
        return False
    if str(payload.get("duplicate_check") or "").strip().lower() != "not dup":
        return False
    if str(payload.get("scrape_status") or "").strip().lower() != "done":
        return False
    return True


def _write_lead_json_to_temp_before_db(
    payload: dict[str, Any],
    *,
    row_id: int,
) -> Path:
    """
    Write the final lead object to ``temp/lead{created_at}.json`` (same token scheme as ``enrichment.py``),
    immediately before persisting that payload to PostgreSQL.
    """
    temp_dir = REPO_ROOT / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    token = lead_generator._lead_created_time_token(payload, fallback_index=row_id)
    base = f"lead{token}"
    path = temp_dir / f"{base}.json"
    if path.exists():
        suffix = 2
        while (temp_dir / f"{base}_{suffix}.json").exists():
            suffix += 1
        path = temp_dir / f"{base}_{suffix}.json"
    public = _lead_public_json_for_export(payload)
    path.write_text(
        json.dumps(public, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    LOGGER.info("Wrote lead JSON before DB payload update: %s", path)
    return path


def _persist_final_scrape_payloads_only(
    dsn: str,
    worker_id: str,
    batch_id: str,
    processed_rows: list[tuple[int, dict[str, Any]]],
) -> None:
    """
    Single point in the pipeline that updates ``payload``: only rows that completed
    Scrapingdog enrichment successfully (``scrape_status == 'done'``).
    """
    rows_to_write = [
        (row_id, payload)
        for row_id, payload in processed_rows
        if _row_should_write_payload_after_scrape(payload)
    ]
    if not rows_to_write:
        return

    for row_id, payload in rows_to_write:
        _write_lead_json_to_temp_before_db(payload, row_id=row_id)

    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                UPDATE data
                SET payload = %s::jsonb,
                    duplicate_check = %s,
                    email_status = %s
                WHERE id = %s
                  AND batch_id = %s::uuid
                  AND claimed_by = %s
                  AND status = 'in_progress'
                """,
                [
                    (
                        json.dumps(
                            _lead_public_json_for_export(payload),
                            default=str,
                        ),
                        _initial_duplicate_check(payload),
                        _initial_email_status(payload),
                        row_id,
                        batch_id,
                        worker_id,
                    )
                    for row_id, payload in rows_to_write
                ],
            )
        conn.commit()


def _complete_processed_rows(
    dsn: str,
    worker_id: str,
    batch_id: str,
    processed_rows: list[tuple[int, dict[str, Any]]],
) -> int:
    if not processed_rows:
        return 0

    with get_connection(dsn) as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                UPDATE data
                SET status = 'done',
                    acked_at = NOW()
                WHERE id = %s
                  AND batch_id = %s::uuid
                  AND claimed_by = %s
                  AND status = 'in_progress'
                """,
                [
                    (row_id, batch_id, worker_id)
                    for row_id, _payload in processed_rows
                ],
            )
            updated_count = cur.rowcount
        conn.commit()
    return int(updated_count or 0)


def process_rows(
    dsn: str,
    worker_id: str,
    batch_id: str,
    rows: list[dict[str, Any]],
) -> int:
    _, completed_count = asyncio.run(
        _process_claimed_rows(dsn, worker_id, batch_id, rows)
    )
    return completed_count


def worker_loop(dsn: str, worker_id: str, batch_size: int = 1000) -> int:
    processed_total = 0

    while True:
        batch = claim_batch(dsn, worker_id, batch_size)
        batch_id = batch["batch_id"]
        rows = batch["rows"]
        if not batch_id:
            LOGGER.info("Worker %s found no more pending rows.", worker_id)
            return processed_total

        completed_early = process_rows(dsn, worker_id, batch_id, rows)
        updated_count = ack_batch(dsn, worker_id, batch_id)
        processed_total += completed_early + updated_count
        LOGGER.info(
            "Worker %s processed batch %s with %s rows.",
            worker_id,
            batch_id,
            completed_early + updated_count,
        )


def start_workers(dsn: str, worker_count: int = 30, batch_size: int = 1000) -> int:
    total_processed = 0
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = [
            executor.submit(worker_loop, dsn, f"worker-{index:02d}", batch_size)
            for index in range(1, worker_count + 1)
        ]
        for future in as_completed(futures):
            total_processed += future.result()

    LOGGER.info(
        "All workers finished. Total acknowledged rows: %s",
        total_processed,
    )
    return total_processed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="PostgreSQL-backed batch provider")
    parser.add_argument(
        "--dsn",
        default=DEFAULT_DSN,
        help="PostgreSQL DSN. Defaults to the built-in local Leads database DSN.",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    subparsers.add_parser("setup-db", help="Create required tables and SQL functions.")

    append_parser = subparsers.add_parser("append-file", help="Append an Excel file into data.")
    append_parser.add_argument("--excel-path", required=True, help="Path to the Excel file.")
    workers_parser = subparsers.add_parser(
        "start-workers",
        help="Start worker threads that claim and acknowledge batches until data is exhausted.",
    )
    workers_parser.add_argument("--worker-count", type=int, default=30)
    workers_parser.add_argument("--batch-size", type=int, default=1000)

    return parser.parse_args()


def main() -> int:
    args = parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    if args.command == "setup-db":
        setup_database(args.dsn)
        return 0

    if args.command == "append-file":
        append_excel_file(args.dsn, Path(args.excel_path))
        return 0

    if args.command == "start-workers":
        start_workers(args.dsn, args.worker_count, args.batch_size)
        return 0

    raise ValueError(f"Unknown command: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())

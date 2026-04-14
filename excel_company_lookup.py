"""
Look up a company row in a local Excel database and return LinkedIn, website, name,
locality (country, US-only state, city), and employee count.

- **Existence**: a row is considered present when ``lookup_value`` matches a cell in the chosen
  lookup column (see ``lookup_mode`` / ``lookup_column``).
- If no matching row exists, ``CompanyLookupResult.exists`` is ``False`` and all string fields are empty.

Configure the workbook path via ``DEFAULT_DATABASE_XLSX`` (or pass ``xlsx_path`` each call).

Typical Apollo / export column names are recognized; headers are matched case-insensitively.

Dependencies: ``openpyxl`` (see requirements.txt).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Set your database file name here (same folder as this script by default)
# ---------------------------------------------------------------------------
DEFAULT_DATABASE_XLSX: Path = Path(__file__).resolve().parent / "database.xlsx"
DEFAULT_SHEET_NAME: Optional[str] = None  # None = first sheet


# Header synonyms (first match wins). All matching is case-insensitive on stripped text.
_COL_COMPANY_LINKEDIN: Sequence[str] = (
    "Company Linkedin Url",
    "Company LinkedIn Url",
    "company_linkedin",
    "company linkedin url",
    "Company LinkedIn URL",
    "Linkedin Url",
    "linkedin company",
)
_COL_WEBSITE: Sequence[str] = (
    "Website",
    "Company Website",
    "website",
    "organization_website",
    "sanitized_organization_website",
    "li_website",
)
_COL_COMPANY_NAME: Sequence[str] = (
    "Company Name",
    "company_name",
    "sanitized_organization_name_unanalyzed",
    "Organization Name",
    "",
)
_COL_CITY: Sequence[str] = (
    "City",
    "Company City",
    "person_location_city",
    "city",
)
_COL_STATE: Sequence[str] = (
    "State",
    "Company State",
    "person_location_state",
    "person_location_state_with_country",
    "state",
)
_COL_COUNTRY: Sequence[str] = (
    "Country",
    "Company Country",
    "person_location_country",
    "country",
    "hq_country",
)
_COL_EMPLOYEES: Sequence[str] = (
    "# Employees",
    "Employees",
    "employee_count",
    "Employee Count",
    "company_employee_count",
    "# employees",
)

# Lookup modes: which column to match ``lookup_value`` against
_LOOKUP_BY_EMAIL: Sequence[str] = (
    "email",
    "Email",
    "person_email",
    "personal_email",
    "person_email_analyzed",
    "work email",
)
_LOOKUP_BY_COMPANY_LINKEDIN: Sequence[str] = _COL_COMPANY_LINKEDIN
_LOOKUP_BY_COMPANY_NAME: Sequence[str] = _COL_COMPANY_NAME


@dataclass
class CompanyLookupResult:
    """Result of a single-row company fetch from the workbook."""

    exists: bool
    company_linkedin: str = ""
    website: str = ""
    company_name: str = ""
    city: str = ""
    state: str = ""  # filled only when country is US
    country: str = ""
    employee_count: str = ""
    matched_row: Optional[int] = None  # 1-based Excel row index, if found
    lookup_column_used: str = ""
    meta: Dict[str, Any] = field(default_factory=dict)


def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_header(h: Any) -> str:
    return _norm_cell(h).lower()


def _normalize_for_match_email(s: str) -> str:
    return _norm_cell(s).lower()


def _normalize_for_match_url(s: str) -> str:
    t = _norm_cell(s).lower().rstrip("/")
    t = re.sub(r"^https?://(www\.)?", "", t)
    return t


def _normalize_for_match_generic(s: str) -> str:
    return _norm_cell(s).lower()


def _header_row_map(ws: Worksheet, max_scan_cols: int = 512) -> Dict[str, int]:
    """Map normalized header text -> 1-based column index."""
    out: Dict[str, int] = {}
    for c in range(1, max_scan_cols + 1):
        h = ws.cell(1, c).value
        if h is None or str(h).strip() == "":
            continue
        key = _norm_header(h)
        if key not in out:
            out[key] = c
    return out


def _resolve_column(
    header_map: Dict[str, int],
    candidates: Sequence[str],
) -> Optional[int]:
    for name in candidates:
        k = _norm_header(name)
        if k in header_map:
            return header_map[k]
    return None


def _is_united_states(country: str) -> bool:
    t = _norm_cell(country).lower()
    return t in (
        "us",
        "usa",
        "u.s.",
        "u.s.a.",
        "united states",
        "united states of america",
        "america",
    )


def _pick_lookup_candidates(lookup_mode: str) -> Sequence[str]:
    m = (lookup_mode or "email").strip().lower()
    if m in ("email", "person_email"):
        return _LOOKUP_BY_EMAIL
    if m in ("company_linkedin", "linkedin", "company linkedin"):
        return _LOOKUP_BY_COMPANY_LINKEDIN
    if m in ("company_name", "company", "name"):
        return _LOOKUP_BY_COMPANY_NAME
    raise ValueError(
        f"unknown lookup_mode={lookup_mode!r}; use email | company_linkedin | company_name"
    )


def _values_equal_for_mode(
    lookup_mode: str,
    cell_text: str,
    lookup_value: str,
) -> bool:
    m = (lookup_mode or "email").strip().lower()
    a, b = _norm_cell(cell_text), _norm_cell(lookup_value)
    if not a or not b:
        return False
    if m in ("email", "person_email"):
        return _normalize_for_match_email(a) == _normalize_for_match_email(b)
    if m in ("company_linkedin", "linkedin", "company linkedin"):
        return _normalize_for_match_url(a) == _normalize_for_match_url(b)
    if m in ("company_name", "company", "name"):
        return _normalize_for_match_generic(a) == _normalize_for_match_generic(b)
    return a.lower() == b.lower()


def check_row_exists_in_excel(
    lookup_value: str,
    *,
    xlsx_path: Optional[Union[str, Path]] = None,
    sheet_name: Optional[str] = None,
    lookup_mode: str = "email",
    lookup_column: Optional[str] = None,
) -> bool:
    """
    Return ``True`` if a row exists where the lookup column equals ``lookup_value``
    (per ``lookup_mode`` or explicit ``lookup_column`` header).
    """
    r = fetch_company_from_excel(
        lookup_value,
        xlsx_path=xlsx_path,
        sheet_name=sheet_name,
        lookup_mode=lookup_mode,
        lookup_column=lookup_column,
    )
    return r.exists


def fetch_company_from_excel(
    lookup_value: str,
    *,
    xlsx_path: Optional[Union[str, Path]] = None,
    sheet_name: Optional[str] = None,
    lookup_mode: str = "email",
    lookup_column: Optional[str] = None,
) -> CompanyLookupResult:
    """
    Find the first row matching ``lookup_value`` and return company fields.

    Parameters
    ----------
    lookup_value
        Value to match (e.g. email, company LinkedIn URL, or company name).
    xlsx_path
        Path to ``.xlsx``. Defaults to ``DEFAULT_DATABASE_XLSX``.
    sheet_name
        Worksheet name; default first sheet or ``DEFAULT_SHEET_NAME`` if set.
    lookup_mode
        ``email`` | ``company_linkedin`` | ``company_name`` — chooses which column group to match.
        Ignored if ``lookup_column`` is set.
    lookup_column
        Exact header text in row 1 to use for matching (overrides ``lookup_mode``).
    """
    path = Path(xlsx_path) if xlsx_path is not None else Path(DEFAULT_DATABASE_XLSX)
    meta: Dict[str, Any] = {"path": str(path)}

    if not path.is_file():
        return CompanyLookupResult(
            exists=False,
            meta={**meta, "error": "file_not_found"},
        )

    lookup_value = _norm_cell(lookup_value)
    if not lookup_value:
        return CompanyLookupResult(
            exists=False,
            meta={**meta, "error": "empty_lookup_value"},
        )

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return CompanyLookupResult(
            exists=False,
            meta={**meta, "error": f"load_workbook_failed: {e}"},
        )

    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return CompanyLookupResult(
                    exists=False,
                    meta={**meta, "error": f"unknown_sheet:{sheet_name}"},
                )
            ws = wb[sheet_name]
        else:
            sn = DEFAULT_SHEET_NAME
            ws = wb[sn] if sn and sn in wb.sheetnames else wb[wb.sheetnames[0]]
            meta["sheet"] = ws.title

        header_map = _header_row_map(ws)
        explicit_lookup = lookup_column is not None and str(lookup_column).strip() != ""
        if explicit_lookup:
            lk_col = _resolve_column(header_map, (str(lookup_column).strip(),))
            used_name = str(lookup_column).strip()
        else:
            candidates = _pick_lookup_candidates(lookup_mode)
            lk_col = _resolve_column(header_map, candidates)
            used_name = ""
            if lk_col is not None:
                mc = int(ws.max_column or 1)
                for c in range(1, mc + 1):
                    if c == lk_col:
                        used_name = _norm_cell(ws.cell(1, c).value)
                        break
        if lk_col is None:
            return CompanyLookupResult(
                exists=False,
                meta={
                    **meta,
                    "error": "lookup_column_not_found",
                    "lookup_mode": lookup_mode,
                },
            )

        meta["lookup_column_used"] = used_name or lookup_column or lookup_mode

        # Data columns
        c_li = _resolve_column(header_map, _COL_COMPANY_LINKEDIN)
        c_web = _resolve_column(header_map, _COL_WEBSITE)
        c_name = _resolve_column(header_map, _COL_COMPANY_NAME)
        c_city = _resolve_column(header_map, _COL_CITY)
        c_state = _resolve_column(header_map, _COL_STATE)
        c_country = _resolve_column(header_map, _COL_COUNTRY)
        c_emp = _resolve_column(header_map, _COL_EMPLOYEES)

        max_row = ws.max_row or 1
        for r in range(2, max_row + 1):
            cell_val = ws.cell(r, lk_col).value
            if explicit_lookup:
                match_ok = (
                    _normalize_for_match_generic(_norm_cell(cell_val))
                    == _normalize_for_match_generic(lookup_value)
                )
            else:
                match_ok = _values_equal_for_mode(
                    lookup_mode, _norm_cell(cell_val), lookup_value
                )
            if match_ok:
                country = _norm_cell(ws.cell(r, c_country).value) if c_country else ""
                state_raw = _norm_cell(ws.cell(r, c_state).value) if c_state else ""
                state_out = state_raw if _is_united_states(country) else ""

                return CompanyLookupResult(
                    exists=True,
                    company_linkedin=_norm_cell(ws.cell(r, c_li).value) if c_li else "",
                    website=_norm_cell(ws.cell(r, c_web).value) if c_web else "",
                    company_name=_norm_cell(ws.cell(r, c_name).value) if c_name else "",
                    city=_norm_cell(ws.cell(r, c_city).value) if c_city else "",
                    state=state_out,
                    country=country,
                    employee_count=_norm_cell(ws.cell(r, c_emp).value) if c_emp else "",
                    matched_row=r,
                    lookup_column_used=meta["lookup_column_used"],
                    meta=meta,
                )

        return CompanyLookupResult(
            exists=False,
            lookup_column_used=str(meta.get("lookup_column_used") or ""),
            meta={**meta, "reason": "no_matching_row"},
        )
    finally:
        wb.close()


def as_dict(result: CompanyLookupResult) -> Dict[str, Any]:
    """Flat dict for JSON / APIs."""
    return {
        "exists": result.exists,
        "company_linkedin": result.company_linkedin,
        "website": result.website,
        "company_name": result.company_name,
        "city": result.city,
        "state": result.state,
        "country": result.country,
        "employee_count": result.employee_count,
        "matched_row": result.matched_row,
        "lookup_column_used": result.lookup_column_used,
        "meta": result.meta,
    }


def main() -> None:
    import argparse
    import json

    p = argparse.ArgumentParser(
        description="Look up a company row in a local Excel DB (see DEFAULT_DATABASE_XLSX)."
    )
    p.add_argument(
        "lookup_value",
        nargs="?",
        default="",
        help="Value to find (email, company LinkedIn URL, or company name)",
    )
    p.add_argument(
        "-i",
        "--input",
        type=Path,
        default=None,
        help=f"Path to .xlsx (default: {DEFAULT_DATABASE_XLSX})",
    )
    p.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")
    p.add_argument(
        "--mode",
        default="email",
        choices=("email", "company_linkedin", "company_name"),
        help="Which column group to match",
    )
    p.add_argument(
        "--column",
        default=None,
        dest="lookup_column",
        metavar="HEADER",
        help="Exact row-1 header to match (overrides --mode)",
    )
    p.add_argument("--json", action="store_true", help="Print JSON")
    args = p.parse_args()

    if not args.lookup_value.strip():
        raise SystemExit("Provide lookup_value (positional) or set it in code.")

    out = fetch_company_from_excel(
        args.lookup_value,
        xlsx_path=args.input,
        sheet_name=args.sheet,
        lookup_mode=args.mode,
        lookup_column=args.lookup_column,
    )
    if args.json:
        print(json.dumps(as_dict(out), indent=2, ensure_ascii=False, default=str))
    else:
        print(f"exists: {out.exists}")
        if out.exists:
            print(f"company_name: {out.company_name}")
            print(f"company_linkedin: {out.company_linkedin}")
            print(f"website: {out.website}")
            print(f"city: {out.city}")
            print(f"state (US only): {out.state}")
            print(f"country: {out.country}")
            print(f"employee_count: {out.employee_count}")
            print(f"row: {out.matched_row}")
        else:
            print("(no row matched)")


if __name__ == "__main__":
    main()
